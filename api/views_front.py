# cython:language_level=3
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect
from django.contrib.auth.hashers import make_password
from django.http import JsonResponse
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.contrib import auth
from api.models import RustDeskPeer, RustDesDevice, UserProfile, ShareLink, ConnLog, FileLog, CustomAppConfig, SupportAgent, AgentConnectionLog, Group, RelayServer, RustDeskToken
from django.forms.models import model_to_dict
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.conf import settings

from itertools import chain
from django.db.models.fields import DateTimeField, DateField, CharField, TextField
import datetime
from django.db.models import Model
import json
import time
import hashlib
import sys

from io import BytesIO
import xlwt
from django.utils.translation import gettext as _
from django.core.cache import cache
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
import random
import uuid
import zipfile
import os
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import base64
import requests

salt = 'xiaomo'
EFFECTIVE_SECONDS = 7200

def send_kakao_password_notification(phone, username, new_password, company_name=''):
    """
    카카오 알림톡을 통해 비밀번호 변경 알림 발송 (일괄 암호변경용)
    
    기존 send_password_change_notification 함수와 동일한 API 사용
    템플릿 코드: MDTL04
    
    Args:
        phone: 수신자 전화번호
        username: 사용자명
        new_password: 새 비밀번호 (현재 템플릿에서는 사용 안함)
        company_name: 회사명 (현재 템플릿에서는 사용 안함)
    
    Returns:
        bool: 발송 성공 여부
    """
    import json as json_module
    
    if not phone:
        return False
    
    # 전화번호 정규화 (하이픈, 공백 제거)
    phone = phone.replace('-', '').replace(' ', '').strip()
    
    # 카카오 알림톡 API URL 확인
    api_url = getattr(settings, 'KAKAO_ALIMTALK_API_URL', '')
    if not api_url:
        print(f"[카카오 알림톡] API URL이 설정되지 않았습니다. (테스트 모드 - 수신: {phone}, 사용자: {username})")
        return True  # 테스트 모드에서는 성공으로 처리
    
    # 카카오 알림톡 API 요청 데이터 구성
    data = {
        "titleinfo": {
            "id": getattr(settings, 'KAKAO_ALIMTALK_ID', ''),
            "pw": getattr(settings, 'KAKAO_ALIMTALK_PW', ''),
            "temNum": getattr(settings, 'KAKAO_ALIMTALK_TEM_NUM', '')
        },
        "messageList": [
            {
                "rcvNum": phone,
                "title": "",
                "btnURL1": "",
                "msg": """원격솔루션 MDesk "비밀번호"가 변경되었습니다.
팀원분들은 관리자에게 변경된 "비밀번호"를 문의 바랍니다.""",
                "ptno": "",
                "rcvDate": "",
                "rcvTime": "",
                "msgtype": "KKO",
                "sendNum": getattr(settings, 'KAKAO_ALIMTALK_SEND_NUM', ''),
                "temCode": "MDTL04"
            }
        ]
    }
    
    try:
        # JSON 문자열로 변환 후 MSG 키에 담아서 POST 전송
        json_data = json_module.dumps(data, ensure_ascii=False)
        print(f"[카카오 알림톡] 발송 시도 - {username}({phone})")
        
        response = requests.post(
            api_url, 
            data={'MSG': json_data},
            headers={'Content-Type': 'application/x-www-form-urlencoded; charset=utf-8'},
            timeout=10
        )
        print(f"[카카오 알림톡] 발송 결과: {username}({phone}) - {response.text}")
        return True
    except Exception as e:
        print(f"[카카오 알림톡] 발송 오류: {str(e)}")
        return False


def getStrMd5(s):
    if not isinstance(s, (str,)):
        s = str(s)

    myHash = hashlib.md5()
    myHash.update(s.encode())

    return myHash.hexdigest()


def model_to_dict2(instance, fields=None, exclude=None, replace=None, default=None):
    """
    :params instance: 模型对象，不能是queryset数据集
    :params fields: 指定要展示的字段数据，('字段1','字段2')
    :params exclude: 指定排除掉的字段数据,('字段1','字段2')
    :params replace: 将字段名字修改成需要的名字，{'数据库字段名':'前端展示名'}
    :params default: 新增不存在的字段数据，{'字段':'数据'}
    """
    # 对传递进来的模型对象校验
    if not isinstance(instance, Model):
        raise Exception(_('model_to_dict는 모델 객체만 받을 수 있습니다'))
    # 对替换数据库字段名字校验
    if replace and type(replace) == dict:   # noqa
        for replace_field in replace.values():
            if hasattr(instance, replace_field):
                raise Exception(_(f'model_to_dict: 변경하려는 {replace_field} 필드가 이미 존재합니다'))
    # 对要新增的默认值进行校验
    if default and type(default) == dict:   # noqa
        for default_key in default.keys():
            if hasattr(instance, default_key):
                raise Exception(_(f'model_to_dict: 기본값을 추가하려 하지만 {default_key} 필드가 이미 존재합니다'))  # noqa
    opts = instance._meta
    data = {}
    for f in chain(opts.concrete_fields, opts.private_fields, opts.many_to_many):
        # 源码下：这块代码会将时间字段剔除掉，我加上一层判断，让其不再剔除时间字段
        if not getattr(f, 'editable', False):
            if type(f) == DateField or type(f) == DateTimeField:   # noqa
                pass
            else:
                continue
        # 如果fields参数传递了，要进行判断
        if fields is not None and f.name not in fields:
            continue
        # 如果exclude 传递了，要进行判断
        if exclude and f.name in exclude:
            continue

        key = f.name
        # 获取字段对应的数据
        if type(f) == DateTimeField:   # noqa
            # 字段类型是，DateTimeFiled 使用自己的方式操作
            value = getattr(instance, key)
            value = datetime.datetime.strftime(value, '%Y-%m-%d %H:%M')
        elif type(f) == DateField:   # noqa
            # 字段类型是，DateFiled 使用自己的方式操作
            value = getattr(instance, key)
            value = datetime.datetime.strftime(value, '%Y-%m-%d')
        elif type(f) == CharField or type(f) == TextField:   # noqa
            # 字符串数据是否可以进行序列化，转成python结构数据
            value = getattr(instance, key)
            try:
                value = json.loads(value)
            except Exception as _:  # noqa
                value = value
        else:  # 其他类型的字段
            # value = getattr(instance, key)
            key = f.name
            value = f.value_from_object(instance)
            # data[f.name] = f.value_from_object(instance)
        # 1、替换字段名字
        if replace and key in replace.keys():
            key = replace.get(key)
        data[key] = value
    # 2、新增默认的字段数据
    if default:
        data.update(default)
    return data


def index(request):
    print('sdf', sys.argv)
    if request.user and request.user.username != 'AnonymousUser':
        return HttpResponseRedirect('/api/work')
    return HttpResponseRedirect('/api/user_action?action=login')


def user_action(request):
    action = request.GET.get('action', '')
    
    # 로그인 관련 액션은 admin.787.kr로만 접근 가능
    if action in ['login', 'login_2fa', 'register', 'logout', 'edit_profile']:
        host = request.get_host()
        admin_host = getattr(settings, 'ID_SERVER', 'admin.787.kr')
        if not (host.startswith(admin_host) or host.startswith('admin.localhost')):
            return HttpResponse(_('접근이 거부되었습니다. {}로 접속해주세요.').format(admin_host), status=403)
    
    if action == 'login':
        return user_login(request)
    elif action == 'login_2fa':
        return user_login_2fa_verify(request)
    elif action == 'register':
        return user_register(request)
    elif action == 'logout':
        return user_logout(request)
    elif action == 'check_phone':
        return check_phone_duplicate(request)
    elif action == 'send_verify':
        return send_verify_code(request)
    elif action == 'verify_phone':
        return verify_phone_code(request)
    elif action == 'find_password':
        return find_password(request)
    elif action == 'reset_password':
        return reset_password(request)
    elif action == 'edit_profile':
        return edit_profile(request)
    else:
        return


def _send_web_2fa_email(email, code, username):
    """웹 로그인 2FA 인증코드 이메일 발송"""
    try:
        subject = _('MDesk 로그인 2차 인증코드')
        message = f"""안녕하세요, {username}님.

MDesk 로그인 2차 인증코드입니다.

인증코드: {code}

이 인증코드는 5분간 유효합니다.
본인이 로그인을 시도하지 않았다면, 비밀번호를 즉시 변경해주세요.

감사합니다.
MDesk 원격지원"""
        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [email], fail_silently=False)
        return True
    except Exception as e:
        print(f"[2FA 이메일] 웹 로그인 발송 실패: {str(e)}")
        return False


def _send_web_2fa_sms(phone, code, username):
    """웹 로그인 2FA 인증코드 카카오 알림톡 발송"""
    try:
        send_kakao_alimtalk(phone, code)
        return True
    except Exception as e:
        print(f"[2FA 알림톡] 웹 로그인 발송 실패: {str(e)}")
        return False


def _generate_2fa_code():
    """6자리 2FA 인증코드 생성"""
    return str(random.randint(100000, 999999))


def user_login(request):
    if request.method == 'GET':
        return render(request, 'login.html')

    username = request.POST.get('account', '')
    password = request.POST.get('password', '')
    if not username or not password:
        return JsonResponse({'code': 0, 'msg': _('문제가 발생했습니다. 사용자명 또는 비밀번호를 가져올 수 없습니다.')})

    user = auth.authenticate(username=username, password=password)
    if not user:
        return JsonResponse({'code': 0, 'msg': _('계정 또는 비밀번호가 틀렸습니다!')})

    # 2차 인증(2FA) 확인
    email_2fa = getattr(user, 'email_2fa', False)
    phone_2fa = getattr(user, 'phone_2fa', False)
    tfa_required = email_2fa or phone_2fa

    if tfa_required:
        code = _generate_2fa_code()
        tfa_key = f'2fa_web_{uuid.uuid4().hex}'
        cache.set(tfa_key, {
            'code': code,
            'user_id': user.id,
            'attempts': 0,
        }, timeout=300)  # 5분

        sent_methods = []
        if email_2fa and user.email and _send_web_2fa_email(user.email, code, user.username):
            email_parts = user.email.split('@')
            sent_methods.append({'type': 'email', 'target': email_parts[0][:2] + '***@' + email_parts[1]})
        if phone_2fa and user.phone and _send_web_2fa_sms(user.phone, code, user.username):
            phone_clean = user.phone.replace('-', '')
            sent_methods.append({'type': 'phone', 'target': phone_clean[:3] + '****' + phone_clean[-4:]})

        if not sent_methods:
            cache.delete(tfa_key)
            auth.login(request, user)
            return JsonResponse({'code': 1, 'url': '/api/work', 'user_pkid': user.id, 'username': user.username})

        return JsonResponse({
            'code': 0,
            'tfa_required': True,
            'tfa_key': tfa_key,
            'tfa_methods': sent_methods,
            'msg': _('2차 인증이 필요합니다. 발송된 인증코드를 입력해주세요.')
        })

    auth.login(request, user)
    return JsonResponse({'code': 1, 'url': '/api/work', 'user_pkid': user.id, 'username': user.username})


def user_login_2fa_verify(request):
    """웹 로그인 2차 인증코드 검증"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': _('잘못된 요청입니다.')})

    tfa_key = request.POST.get('tfa_key', '')
    tfa_code = request.POST.get('tfa_code', '')
    if not tfa_key or not tfa_code:
        return JsonResponse({'code': 0, 'msg': _('인증코드를 입력해주세요.')})

    tfa_data = cache.get(tfa_key)
    if not tfa_data:
        return JsonResponse({'code': 0, 'msg': _('인증코드가 만료되었습니다. 다시 로그인해주세요.')})

    if tfa_data.get('attempts', 0) >= 5:
        cache.delete(tfa_key)
        return JsonResponse({'code': 0, 'msg': _('인증 시도 횟수를 초과했습니다. 다시 로그인해주세요.')})

    if tfa_code.strip() != tfa_data['code']:
        tfa_data['attempts'] = tfa_data.get('attempts', 0) + 1
        cache.set(tfa_key, tfa_data, timeout=300)
        remaining = 5 - tfa_data['attempts']
        return JsonResponse({'code': 0, 'msg': _('인증코드가 일치하지 않습니다. (남은 시도: {}회)').format(remaining)})

    cache.delete(tfa_key)
    user = UserProfile.objects.filter(id=tfa_data['user_id']).first()
    if not user:
        return JsonResponse({'code': 0, 'msg': _('사용자를 찾을 수 없습니다.')})

    auth.login(request, user)
    return JsonResponse({'code': 1, 'url': '/api/work', 'user_pkid': user.id, 'username': user.username})


def user_register(request):
    info = ''
    if request.method == 'GET':
        return render(request, 'reg.html')
    ALLOW_REGISTRATION = settings.ALLOW_REGISTRATION
    result = {
        'code': 0,
        'msg': ''
    }
    if not ALLOW_REGISTRATION:
        result['msg'] = _('현재 회원가입이 불가합니다. 관리자에게 문의하세요!')
        return JsonResponse(result)

    username = request.POST.get('user', '')
    password1 = request.POST.get('pwd', '')
    company_name = request.POST.get('company', '')
    email = request.POST.get('email', '')
    phone = request.POST.get('phone', '')
    phone_verified = request.POST.get('phone_verified', '0') == '1'

    if len(username) <= 3:
        info = _('사용자명은 3자 이상이어야 합니다')
        result['msg'] = info
        return JsonResponse(result)

    if len(password1) < 8 or len(password1) > 20:
        info = _('비밀번호 길이가 요구사항에 맞지 않습니다. 8~20자여야 합니다.')
        result['msg'] = info
        return JsonResponse(result)

    # 이메일 중복 확인
    if email:
        existing_email = UserProfile.objects.filter(Q(email=email) & ~Q(email='')).first()
        if existing_email:
            info = _('이미 사용 중인 이메일입니다.')
            result['msg'] = info
            return JsonResponse(result)

    # 휴대전화 중복 확인
    if phone:
        existing_phone = UserProfile.objects.filter(Q(phone=phone) & ~Q(phone='')).first()
        if existing_phone:
            info = _('이미 사용 중인 휴대전화번호입니다.')
            result['msg'] = info
            return JsonResponse(result)

    user = UserProfile.objects.filter(Q(username=username)).first()
    if user:
        info = _('사용자명이 이미 존재합니다.')
        result['msg'] = info
        return JsonResponse(result)
    
    user = UserProfile(
        username=username,
        password=make_password(password1),
        company_name=company_name,
        email=email,
        phone=phone,
        phone_verified=phone_verified,
        is_admin=True if UserProfile.objects.count() == 0 else False,
        is_superuser=True if UserProfile.objects.count() == 0 else False,
        is_active=True
    )
    user.save()
    result['msg'] = info
    result['code'] = 1
    return JsonResponse(result)


def send_email_verify_code(email, verify_code):
    """이메일로 인증번호 발송"""
    try:
        subject = _('인증번호 발송')
        message = f"""
안녕하세요.

요청하신 인증번호는 다음과 같습니다:

인증번호: {verify_code}

이 인증번호는 5분간 유효합니다.
타인에게 노출되지 않도록 주의해주세요.

감사합니다.
        """.strip()
        
        from_email = settings.DEFAULT_FROM_EMAIL
        recipient_list = [email]
        
        send_mail(
            subject,
            message,
            from_email,
            recipient_list,
            fail_silently=False,
        )
        return True
    except Exception as e:
        print(f'이메일 발송 오류: {str(e)}')
        return False


def send_kakao_alimtalk(phone, verify_code):
    """카카오 알림톡으로 인증번호 발송"""
    current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        # 전화번호에서 하이픈 제거
        phone_clean = phone.replace('-', '')
        
        print('=' * 60)
        print(f'[카카오 알림톡 발송 시작] {current_time}')
        print(f'  - 수신번호: {phone} (정제: {phone_clean})')
        print(f'  - 인증번호: {verify_code}')
        print(f'  - API URL: {settings.KAKAO_ALIMTALK_API_URL}')
        print(f'  - 발신번호: {settings.KAKAO_ALIMTALK_SEND_NUM}')
        print(f'  - 템플릿 코드: MDTL03 (인증번호)')
        
        # 카카오 알림톡 API 요청 데이터 구성
        data = {
            "titleinfo": {
                "id": settings.KAKAO_ALIMTALK_ID,
                "pw": settings.KAKAO_ALIMTALK_PW,
                "temNum": settings.KAKAO_ALIMTALK_TEM_NUM
            },
            "messageList": [
                {
                    "rcvNum": phone_clean,
                    "title": verify_code,
                    "btnURL1": "",
                    "msg": """■ MDesk 원격지원 안내 ■
본인확인 인증번호 입니다.
타인에게 노출되지 않도록 주의해 주세요.""",
                    "ptno": "",
                    "rcvDate": "",
                    "rcvTime": "",
                    "msgtype": "KKO",
                    "sendNum": settings.KAKAO_ALIMTALK_SEND_NUM,
                    "temCode": "MDTL03"  # 인증번호 발송 템플릿
                }
            ]
        }
        
        # JSON 문자열로 변환
        json_data = json.dumps(data, ensure_ascii=False)
        print(f'  - 요청 데이터 준비 완료')
        print(f'  - 요청 JSON 데이터: {json_data}')
        
        # x-www-form-urlencoded 형식으로 MSG 키에 JSON 문자열 담아서 전송
        form_data = {
            'MSG': json_data
        }
        print(f'  - Form 데이터: MSG={json_data[:100]}...')
        
        # API 호출 - x-www-form-urlencoded 형식으로 전송
        response = requests.post(
            settings.KAKAO_ALIMTALK_API_URL,
            data=form_data,
            headers={
                'Content-Type': 'application/x-www-form-urlencoded; charset=utf-8'
            },
            timeout=10
        )
        
        print(f'  - API 응답 상태 코드: {response.status_code}')
        
        # 응답 확인
        if response.status_code == 200:
            try:
                result_data = response.json()
            except Exception as json_error:
                print(f'[카카오 알림톡 발송 실패] {current_time}')
                print(f'  - JSON 파싱 오류: {str(json_error)}')
                print(f'  - 응답 텍스트: {response.text[:500]}')
                print('=' * 60)
                return False
            
            print(f'  - API 응답 데이터: {result_data}')
            
            # API 응답 구조에 따라 성공 여부 확인
            if result_data.get('RESULT') and len(result_data['RESULT']) > 0:
                result_item = result_data['RESULT'][0]
                if result_item.get('VALUE') == '0':
                    # VALUE가 0이면 에러
                    error_msg = result_item.get('COMMENT', '알 수 없는 오류')
                    print(f'[카카오 알림톡 발송 실패] {current_time}')
                    print(f'  - 오류 메시지: {error_msg}')
                    print(f'  - 전체 응답: {result_data}')
                    print('=' * 60)
                    return False
                else:
                    print(f'[카카오 알림톡 발송 성공] {current_time}')
                    print(f'  - 전화번호: {phone}')
                    print(f'  - 인증번호: {verify_code}')
                    print(f'  - 응답 VALUE: {result_item.get("VALUE")}')
                    print('=' * 60)
                    return True
            else:
                print(f'[카카오 알림톡 발송 실패] {current_time}')
                print(f'  - 응답 형식 오류: {result_data}')
                print(f'  - RESULT 키가 없거나 비어있음')
                print('=' * 60)
                return False
        else:
            print(f'[카카오 알림톡 발송 실패] {current_time}')
            print(f'  - HTTP 상태 코드: {response.status_code}')
            print(f'  - 응답 내용: {response.text[:500]}')
            print('=' * 60)
            return False
            
    except Exception as e:
        print(f'[카카오 알림톡 발송 오류] {current_time}')
        print(f'  - 예외 발생: {str(e)}')
        print(f'  - 예외 타입: {type(e).__name__}')
        import traceback
        print(f'  - 상세 오류:\n{traceback.format_exc()}')
        print('=' * 60)
        return False


def send_reset_password_email(email, reset_token, reset_url):
    """비밀번호 재설정 이메일 발송"""
    try:
        subject = 'MDesk 원격지원 - 비밀번호 재설정'
        message = f"""
안녕하세요.

비밀번호 재설정을 요청하셨습니다.
아래 링크를 클릭하여 비밀번호를 재설정하실 수 있습니다:

{reset_url}

이 링크는 30분간 유효합니다.
만약 비밀번호 재설정을 요청하지 않으셨다면, 이 이메일을 무시하셔도 됩니다.

감사합니다.
        """.strip()
        
        from_email = settings.DEFAULT_FROM_EMAIL
        recipient_list = [email]
        
        send_mail(
            subject,
            message,
            from_email,
            recipient_list,
            fail_silently=False,
        )
        return True
    except Exception as e:
        print(f'이메일 발송 오류: {str(e)}')
        return False


def check_phone_duplicate(request):
    """전화번호 중복 확인"""
    result = {
        'code': 1,
        'msg': ''
    }
    
    if request.method != 'POST':
        result['code'] = 0
        result['msg'] = _('잘못된 요청입니다.')
        return JsonResponse(result)
    
    phone = request.POST.get('phone', '')
    if not phone:
        result['code'] = 0
        result['msg'] = _('전화번호를 입력해주세요.')
        return JsonResponse(result)
    
    # 하이픈 제거
    phone_clean = phone.replace('-', '')
    
    # 기존 사용자 중 해당 전화번호가 있는지 확인
    existing_user = UserProfile.objects.filter(phone=phone_clean).first()
    if existing_user:
        result['code'] = 0
        result['msg'] = _('이미 등록된 전화번호입니다.')
        return JsonResponse(result)
    
    # 하이픈 포함 번호로도 확인
    existing_user = UserProfile.objects.filter(phone=phone).first()
    if existing_user:
        result['code'] = 0
        result['msg'] = _('이미 등록된 전화번호입니다.')
        return JsonResponse(result)
    
    result['msg'] = _('사용 가능한 전화번호입니다.')
    return JsonResponse(result)


def send_verify_code(request):
    """인증번호 발송 (이메일 또는 SMS)"""
    current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    result = {
        'code': 0,
        'msg': ''
    }
    
    print('=' * 60)
    print(f'[인증번호 발송 요청 시작] {current_time}')
    print(f'  - 요청 방식: {request.method}')
    print(f'  - 요청 IP: {request.META.get("REMOTE_ADDR", "알 수 없음")}')
    
    if request.method != 'POST':
        result['msg'] = _('잘못된 요청입니다.')
        print(f'  - 오류: POST 방식이 아님')
        print('=' * 60)
        return JsonResponse(result)
    
    phone = request.POST.get('phone', '')
    email = request.POST.get('email', '')
    username = request.POST.get('username', '')
    
    print(f'  - 전화번호: {phone if phone else "(없음)"}')
    print(f'  - 이메일: {email if email else "(없음)"}')
    print(f'  - 사용자명: {username if username else "(없음)"}')
    
    # 사용자명이 제공된 경우 (비밀번호 찾기 등), 정보가 일치하는지 먼저 확인
    if username:
        user = UserProfile.objects.filter(username=username).first()
        if not user:
            result['msg'] = _('일치하는 사용자 정보를 찾을 수 없습니다.')
            print(f'  - 오류: 사용자 없음 ({username})')
            print('=' * 60)
            return JsonResponse(result)
        
        if email and user.email != email:
            result['msg'] = _('사용자명과 이메일 정보가 일치하지 않습니다.')
            print(f'  - 오류: 이메일 불일치')
            print('=' * 60)
            return JsonResponse(result)
            
        if phone and user.phone != phone:
            result['msg'] = _('사용자명과 휴대전화번호 정보가 일치하지 않습니다.')
            print(f'  - 오류: 전화번호 불일치')
            print('=' * 60)
            return JsonResponse(result)
    
    if not phone and not email:
        result['msg'] = _('휴대전화번호 또는 이메일을 입력해주세요.')
        print(f'  - 오류: 전화번호와 이메일 모두 없음')
        print('=' * 60)
        return JsonResponse(result)
    
    # 인증번호 생성
    verify_code = str(random.randint(100000, 999999))
    print(f'  - 생성된 인증번호: {verify_code}')
    
    # 이메일로 발송
    if email:
        print(f'[이메일 인증번호 발송 시작] {current_time}')
        # 캐시에 인증번호 저장 (5분 유효)
        cache_key = f'verify_code_email_{email}'
        cache.set(cache_key, verify_code, 300)  # 5분
        print(f'  - 캐시 키: {cache_key}')
        print(f'  - 유효시간: 5분 (300초)')
        
        # 이메일 발송
        email_sent = send_email_verify_code(email, verify_code)
        
        if email_sent:
            result['code'] = 1
            result['msg'] = _('인증번호가 이메일로 발송되었습니다.')
            print(f'[이메일 인증번호 발송 성공] {current_time}')
        else:
            result['msg'] = _('이메일 발송에 실패했습니다. 이메일 설정을 확인해주세요.')
            print(f'[이메일 인증번호 발송 실패] {current_time}')
        
        print(f'  - 이메일: {email}')
        print(f'  - 인증번호: {verify_code}')
        print('=' * 60)
        
        return JsonResponse(result)
    
    # 휴대전화로 발송 (카카오 알림톡)
    if phone:
        print(f'[카카오 알림톡 인증번호 발송 시작] {current_time}')
        # 캐시에 인증번호 저장 (5분 유효)
        cache_key = f'verify_code_{phone}'
        cache.set(cache_key, verify_code, 300)  # 5분
        print(f'  - 캐시 키: {cache_key}')
        print(f'  - 유효시간: 5분 (300초)')
        
        # 카카오 알림톡 발송
        alimtalk_sent = send_kakao_alimtalk(phone, verify_code)
        
        if alimtalk_sent:
            result['code'] = 1
            result['msg'] = _('인증번호가 카카오 알림톡으로 발송되었습니다.')
            print(f'[카카오 알림톡 인증번호 발송 완료] {current_time}')
        else:
            # 발송 실패 시에도 콘솔에 출력 (개발/테스트용)
            print(f'[카카오 알림톡 인증번호 발송 실패] {current_time}')
            print(f'  - 휴대전화번호: {phone}')
            print(f'  - 인증번호: {verify_code} (콘솔 확인용)')
            print(f'  - 유효시간: 5분')
            print(f'  - 참고: 서버 콘솔 로그에서 상세 오류 확인 가능')
            
            result['code'] = 1  # 인증번호는 생성되었으므로 code는 1로 설정
            result['msg'] = _('인증번호 발송에 실패했습니다. 관리자에게 문의하세요. (서버 콘솔 로그 확인)')
            result['debug_info'] = {
                'phone': phone,
                'verify_code': verify_code,
                'note': '서버 콘솔에서 상세 오류 로그를 확인하세요.'
            }
        
        print('=' * 60)
        return JsonResponse(result)
    
    result['msg'] = _('휴대전화번호 또는 이메일을 입력해주세요.')
    print(f'  - 오류: 처리할 수 있는 항목 없음')
    print('=' * 60)
    return JsonResponse(result)


def verify_phone_code(request):
    """인증번호 확인 (휴대전화 또는 이메일)"""
    current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    result = {
        'code': 0,
        'msg': ''
    }
    
    print('=' * 60)
    print(f'[인증번호 확인 요청 시작] {current_time}')
    print(f'  - 요청 방식: {request.method}')
    print(f'  - 요청 IP: {request.META.get("REMOTE_ADDR", "알 수 없음")}')
    
    if request.method != 'POST':
        result['msg'] = _('잘못된 요청입니다.')
        print(f'  - 오류: POST 방식이 아님')
        print('=' * 60)
        return JsonResponse(result)
    
    phone = request.POST.get('phone', '')
    email = request.POST.get('email', '')
    code = request.POST.get('code', '')
    
    print(f'  - 전화번호: {phone if phone else "(없음)"}')
    print(f'  - 이메일: {email if email else "(없음)"}')
    print(f'  - 입력된 인증번호: {code if code else "(없음)"}')
    
    if not code:
        result['msg'] = _('인증번호를 입력해주세요.')
        print(f'  - 오류: 인증번호 없음')
        print('=' * 60)
        return JsonResponse(result)
    
    if not phone and not email:
        result['msg'] = _('휴대전화번호 또는 이메일을 입력해주세요.')
        print(f'  - 오류: 전화번호와 이메일 모두 없음')
        print('=' * 60)
        return JsonResponse(result)
    
    # 이메일 인증 확인
    if email:
        print(f'[이메일 인증번호 확인 시작] {current_time}')
        cache_key = f'verify_code_email_{email}'
        stored_code = cache.get(cache_key)
        print(f'  - 캐시 키: {cache_key}')
        print(f'  - 저장된 인증번호: {stored_code if stored_code else "(없음/만료됨)"}')
        print(f'  - 입력된 인증번호: {code}')
        
        if not stored_code:
            result['msg'] = _('인증번호가 만료되었습니다. 다시 발송해주세요.')
            print(f'  - 결과: 실패 (인증번호 만료)')
            print('=' * 60)
            return JsonResponse(result)
        
        if stored_code != code:
            result['msg'] = _('인증번호가 일치하지 않습니다.')
            print(f'  - 결과: 실패 (인증번호 불일치)')
            print('=' * 60)
            return JsonResponse(result)
        
        # 인증 성공 - 캐시에 인증 완료 표시
        cache.set(f'email_verified_{email}', True, 600)  # 10분 유효
        cache.delete(cache_key)  # 인증번호 삭제
        
        result['code'] = 1
        result['msg'] = _('이메일 인증이 완료되었습니다.')
        print(f'  - 결과: 성공 (인증 완료)')
        print(f'  - 인증 완료 플래그 저장 (10분 유효)')
        print('=' * 60)
        return JsonResponse(result)
    
    # 휴대전화 인증 확인
    if phone:
        print(f'[휴대전화 인증번호 확인 시작] {current_time}')
        cache_key = f'verify_code_{phone}'
        stored_code = cache.get(cache_key)
        print(f'  - 캐시 키: {cache_key}')
        print(f'  - 저장된 인증번호: {stored_code if stored_code else "(없음/만료됨)"}')
        print(f'  - 입력된 인증번호: {code}')
        
        if not stored_code:
            result['msg'] = _('인증번호가 만료되었습니다. 다시 발송해주세요.')
            print(f'  - 결과: 실패 (인증번호 만료)')
            print('=' * 60)
            return JsonResponse(result)
        
        if stored_code != code:
            result['msg'] = _('인증번호가 일치하지 않습니다.')
            print(f'  - 결과: 실패 (인증번호 불일치)')
            print('=' * 60)
            return JsonResponse(result)
        
        # 인증 성공 - 캐시에 인증 완료 표시
        cache.set(f'phone_verified_{phone}', True, 600)  # 10분 유효
        cache.delete(cache_key)  # 인증번호 삭제
        
        result['code'] = 1
        result['msg'] = _('휴대전화 인증이 완료되었습니다.')
        print(f'  - 결과: 성공 (인증 완료)')
        print(f'  - 인증 완료 플래그 저장 (10분 유효)')
        print('=' * 60)
        return JsonResponse(result)
    
    result['msg'] = _('휴대전화번호 또는 이메일을 입력해주세요.')
    print(f'  - 오류: 처리할 수 있는 항목 없음')
    print('=' * 60)
    return JsonResponse(result)


def find_password(request):
    """비밀번호 찾기"""
    result = {
        'code': 0,
        'msg': ''
    }
    
    if request.method == 'GET':
        return render(request, 'find_password.html')
    
    if request.method != 'POST':
        result['msg'] = _('잘못된 요청입니다.')
        return JsonResponse(result)
    
    username = request.POST.get('username', '').strip()
    email = request.POST.get('email', '')
    phone = request.POST.get('phone', '')
    
    if not username:
        result['msg'] = _('사용자명(ID)을 입력해주세요.')
        return JsonResponse(result)
    
    user = None
    
    # 사용자명(ID)으로 먼저 찾기
    user = UserProfile.objects.filter(username=username).first()
    if not user:
        result['msg'] = _('일치하는 사용자 정보를 찾을 수 없습니다.')
        return JsonResponse(result)
    
    # 이메일로 찾기
    if email:
        if user.email != email:
            result['msg'] = _('사용자명과 이메일 정보가 일치하지 않습니다.')
            return JsonResponse(result)
    
    # 휴대전화로 찾기
    elif phone:
        # 휴대전화 인증 확인
        cache_key = f'phone_verified_{phone}'
        if not cache.get(cache_key):
            result['msg'] = _('휴대전화 인증을 먼저 완료해주세요.')
            return JsonResponse(result)
        
        if user.phone != phone:
            result['msg'] = _('사용자명과 휴대전화번호 정보가 일치하지 않습니다.')
            return JsonResponse(result)
    else:
        result['msg'] = _('이메일 또는 휴대전화번호를 입력해주세요.')
        return JsonResponse(result)
    
    # 비밀번호 재설정 토큰 생성
    reset_token = getStrMd5(str(time.time()) + salt + user.username)
    
    # 토큰을 캐시에 저장 (30분 유효)
    cache_key = f'reset_token_{reset_token}'
    cache.set(cache_key, user.id, 1800)  # 30분
    
    # 재설정 링크 생성 (ID_SERVER 사용)
    admin_host = settings.ID_SERVER or request.get_host()
    reset_url = f'https://{admin_host}/api/user_action?action=reset_password&token={reset_token}'
    
    # 이메일로 발송
    if email:
        email_sent = send_reset_password_email(user.email, reset_token, reset_url)
        
        if email_sent:
            result['code'] = 1
            result['msg'] = _('비밀번호 재설정 링크가 이메일로 발송되었습니다.')
        else:
            result['msg'] = _('이메일 발송에 실패했습니다. 이메일 설정을 확인해주세요.')
            # 콘솔에도 출력 (개발/테스트용)
            print('=' * 50)
            print(f'[비밀번호 찾기 - 이메일 발송 실패]')
            print(f'사용자명: {user.username}')
            print(f'이메일: {email}')
            print(f'재설정 토큰: {reset_token}')
            print(f'재설정 링크: {reset_url}')
            print('=' * 50)
        
        return JsonResponse(result)
    
    # 휴대전화로 발송 (SMS)
    if phone:
        # 콘솔에 출력 (개발/테스트용 - 실제 SMS 서비스 연동 필요)
        print('=' * 50)
        print(f'[비밀번호 찾기 - SMS]')
        print(f'사용자명: {user.username}')
        print(f'휴대전화: {phone}')
        print(f'재설정 토큰: {reset_token}')
        print(f'재설정 링크: {reset_url}')
        print('=' * 50)
        
        # TODO: 실제 SMS 발송 로직 추가
        # send_reset_sms(phone, reset_token)
        
        result['code'] = 1
        result['msg'] = _('비밀번호 재설정 링크가 발송되었습니다. (콘솔 확인)')
        result['reset_token'] = reset_token
        return JsonResponse(result)
    
    result['msg'] = _('이메일 또는 휴대전화번호를 입력해주세요.')
    return JsonResponse(result)


def reset_password(request):
    """비밀번호 재설정"""
    result = {
        'code': 0,
        'msg': ''
    }
    
    if request.method == 'GET':
        token = request.GET.get('token', '')
        if not token:
            result['msg'] = _('유효하지 않은 링크입니다.')
            return render(request, 'msg.html', {'title': _('오류'), 'msg': result['msg']})
        
        # 토큰 유효성 확인
        cache_key = f'reset_token_{token}'
        user_id = cache.get(cache_key)
        if not user_id:
            result['msg'] = _('만료되었거나 유효하지 않은 링크입니다.')
            return render(request, 'msg.html', {'title': _('오류'), 'msg': result['msg']})
        
        return render(request, 'reset_password.html')
    
    if request.method != 'POST':
        result['msg'] = _('잘못된 요청입니다.')
        return JsonResponse(result)
    
    token = request.POST.get('token', '')
    password = request.POST.get('password', '')
    
    if not token or not password:
        result['msg'] = _('토큰과 비밀번호를 입력해주세요.')
        return JsonResponse(result)
    
    # 토큰 유효성 확인
    cache_key = f'reset_token_{token}'
    user_id = cache.get(cache_key)
    if not user_id:
        result['msg'] = _('만료되었거나 유효하지 않은 토큰입니다.')
        return JsonResponse(result)
    
    # 비밀번호 길이 확인
    if len(password) < 8 or len(password) > 20:
        result['msg'] = _('비밀번호 길이가 요구사항에 맞지 않습니다. 8~20자여야 합니다.')
        return JsonResponse(result)
    
    # 사용자 찾기
    try:
        user = UserProfile.objects.get(id=user_id)
    except UserProfile.DoesNotExist:
        result['msg'] = _('사용자를 찾을 수 없습니다.')
        return JsonResponse(result)
    
    # 비밀번호 변경
    user.set_password(password)
    user.save()
    
    # 기존 로그인 토큰 무효화 (강제 로그아웃)
    from api.models import RustDeskToken
    deleted_tokens = RustDeskToken.objects.filter(uid=user.id).delete()
    
    # 재설정 토큰 삭제
    cache.delete(cache_key)
    
    result['code'] = 1
    result['msg'] = _('비밀번호가 성공적으로 변경되었습니다. 로그인 페이지에서 로그인하세요.')
    return JsonResponse(result)


@login_required(login_url='/api/user_action?action=login')
def edit_profile(request):
    """내정보 수정"""
    username = request.user
    u = UserProfile.objects.get(username=username)
    
    if request.method == 'GET':
        # 사용자의 그룹 정보 조회
        user_group = None
        if u.group_id:
            user_group = Group.objects.filter(id=u.group_id, is_deleted=False).first()
        # 팀원 목록 조회
        from api.models import TeamMember
        team_members = TeamMember.objects.filter(user=u, is_active=True)
        return render(request, 'edit_profile.html', {'u': u, 'user_group': user_group, 'team_members': team_members})
    
    result = {
        'code': 0,
        'msg': ''
    }
    
    company_name = request.POST.get('company_name', '')
    email = request.POST.get('email', '')
    phone = request.POST.get('phone', '')
    phone_verified = request.POST.get('phone_verified', '0') == '1'
    new_password = request.POST.get('new_password', '')
    
    # 이메일 중복 확인 (다른 사용자)
    if email:
        existing_email = UserProfile.objects.filter(Q(email=email) & ~Q(email='') & ~Q(id=u.id)).first()
        if existing_email:
            result['msg'] = _('이미 사용 중인 이메일입니다.')
            return JsonResponse(result)
    
    # 휴대전화 중복 확인 (다른 사용자)
    if phone:
        existing_phone = UserProfile.objects.filter(Q(phone=phone) & ~Q(phone='') & ~Q(id=u.id)).first()
        if existing_phone:
            result['msg'] = _('이미 사용 중인 휴대전화번호입니다.')
            return JsonResponse(result)
    
    # 정보 업데이트
    u.company_name = company_name
    u.email = email
    u.phone = phone
    if phone_verified:
        u.phone_verified = True
    
    # 비밀번호 변경
    password_changed = False
    if new_password:
        if len(new_password) < 8 or len(new_password) > 20:
            result['msg'] = _('비밀번호 길이가 요구사항에 맞지 않습니다. 8~20자여야 합니다.')
            return JsonResponse(result)
        u.set_password(new_password)
        password_changed = True
    
    u.save()
    
    # 비밀번호 변경 시 기존 로그인 토큰 무효화 (강제 로그아웃)
    if password_changed:
        from api.models import RustDeskToken
        RustDeskToken.objects.filter(uid=u.id).delete()
    
    # 비밀번호 변경 시 팀원에게 알림톡 발송
    notified_count = 0
    notify_team = request.POST.get('notify_team', '0') == '1'
    if password_changed and notify_team:
        from api.models import TeamMember
        team_members = TeamMember.objects.filter(user=u, is_active=True)
        
        for member in team_members:
            if member.phone:
                try:
                    # 카카오 알림톡 발송
                    send_password_change_notification(u.username, member.name, member.phone, new_password)
                    notified_count += 1
                except Exception as e:
                    print(f"알림톡 발송 실패: {member.name} - {str(e)}")
    
    # 2차 인증(2FA) 설정 저장 - 둘 중 하나만 선택 가능
    email_2fa = request.POST.get('email_2fa', '0') == '1' and bool(u.email)
    phone_2fa = request.POST.get('phone_2fa', '0') == '1' and bool(u.phone)
    if email_2fa and phone_2fa:
        email_2fa = False  # 상호 배타: 휴대전화 우선
    
    u.email_2fa = email_2fa
    u.phone_2fa = phone_2fa
    u.save()
    
    result['code'] = 1
    result['msg'] = _('정보가 성공적으로 수정되었습니다.')
    result['notified_count'] = notified_count
    return JsonResponse(result)


def send_password_change_notification(username, member_name, phone, new_password):
    """팀원에게 비밀번호 변경 알림톡 발송
    
    템플릿 코드: MDTL04
    메시지 내용:
    원격솔루션 MDesk "비밀번호"가 변경되었습니다.
    팀원분들은 관리자에게 변경된 "비밀번호"를 문의 바랍니다.
    """
    import requests
    import json
    
    # 전화번호 정규화 (하이픈 제거)
    phone = phone.replace('-', '').replace(' ', '')
    
    # 카카오 알림톡 API 호출
    api_url = settings.KAKAO_ALIMTALK_API_URL
    if not api_url:
        print("카카오 알림톡 API URL이 설정되지 않았습니다.")
        return False
    
    # 카카오 알림톡 API 요청 데이터 구성
    data = {
        "titleinfo": {
            "id": settings.KAKAO_ALIMTALK_ID,
            "pw": settings.KAKAO_ALIMTALK_PW,
            "temNum": settings.KAKAO_ALIMTALK_TEM_NUM
        },
        "messageList": [
            {
                "rcvNum": phone,
                "title": "",
                "btnURL1": "",
                "msg": """원격솔루션 MDesk "비밀번호"가 변경되었습니다.
팀원분들은 관리자에게 변경된 "비밀번호"를 문의 바랍니다.""",
                "ptno": "",
                "rcvDate": "",
                "rcvTime": "",
                "msgtype": "KKO",
                "sendNum": settings.KAKAO_ALIMTALK_SEND_NUM,
                "temCode": "MDTL04"
            }
        ]
    }
    
    try:
        # JSON 문자열로 변환 후 MSG 키에 담아서 POST 전송
        json_data = json.dumps(data, ensure_ascii=False)
        print(f"  - 요청 JSON 데이터: {json_data}")
        
        response = requests.post(
            api_url, 
            data={'MSG': json_data},
            headers={'Content-Type': 'application/x-www-form-urlencoded; charset=utf-8'},
            timeout=10
        )
        print(f"알림톡 발송 결과: {member_name}({phone}) - {response.text}")
        return True
    except Exception as e:
        print(f"알림톡 발송 오류: {str(e)}")
        return False


@login_required(login_url='/api/user_action?action=login')
def user_logout(request):
    # info=''
    auth.logout(request)
    return HttpResponseRedirect('/api/user_action?action=login')


def get_recent_remotes_for_user(uid, limit=10):
    """사용자가 최근 접속한 원격지 목록 (ConnLog 기준)"""
    user_peers = RustDeskPeer.objects.filter(Q(uid=uid))
    user_rids = [str(p.rid) for p in user_peers]
    if not user_rids:
        return []
    # 사용자가 from_id(내 기기)로 접속한 경우: from_id IN user_rids, rid=원격지
    logs = list(ConnLog.objects.filter(
        from_id__in=user_rids,
        conn_start__isnull=False
    ).order_by('-conn_start')[:limit * 5])
    seen = set()
    result = []
    rids_from_logs = list(set(x.rid for x in logs if x.rid))
    peer_map = {p.rid: p for p in RustDeskPeer.objects.filter(rid__in=rids_from_logs)}
    device_map = {d.rid: d for d in RustDesDevice.objects.filter(rid__in=rids_from_logs)}
    for log in logs:
        if log.rid in seen or len(result) >= limit:
            continue
        seen.add(log.rid)
        peer = peer_map.get(log.rid)
        device = device_map.get(log.rid)
        alias = peer.alias if peer else (device.hostname if device else log.rid)
        hostname = device.hostname if device else '-'
        result.append({
            'rid': log.rid,
            'alias': alias or log.rid,
            'hostname': hostname,
            'last_conn': log.conn_start,
        })
    return result


def get_connection_stats_for_user(uid):
    """사용자 연결 통계 (오늘, 이번 주, 전체)"""
    user_peers = RustDeskPeer.objects.filter(Q(uid=uid))
    user_rids = [str(p.rid) for p in user_peers]
    if not user_rids:
        return {'today': 0, 'week': 0, 'total': 0}
    now = datetime.datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - datetime.timedelta(days=now.weekday())
    today_count = ConnLog.objects.filter(from_id__in=user_rids, conn_start__gte=today_start).count()
    week_count = ConnLog.objects.filter(from_id__in=user_rids, conn_start__gte=week_start).count()
    total_count = ConnLog.objects.filter(from_id__in=user_rids).count()
    return {'today': today_count, 'week': week_count, 'total': total_count}


def get_single_info(uid):
    peers = RustDeskPeer.objects.filter(Q(uid=uid))
    rids = [x.rid for x in peers]
    peers = {x.rid: model_to_dict(x) for x in peers}
    # print(peers)
    devices = RustDesDevice.objects.filter(rid__in=rids)
    devices = {x.rid: x for x in devices}
    now = datetime.datetime.now()
    for rid, device in devices.items():
        peers[rid]['create_time'] = device.create_time.strftime('%Y-%m-%d')
        peers[rid]['update_time'] = device.update_time.strftime('%Y-%m-%d %H:%M')
        peers[rid]['version'] = device.version
        peers[rid]['memory'] = device.memory
        peers[rid]['cpu'] = device.cpu
        peers[rid]['os'] = device.os
        peers[rid]['status'] = _('온라인') if (now - device.update_time).seconds <= 120 else _('오프라인')

    for rid in peers.keys():
        peers[rid]['has_rhash'] = _('예') if len(peers[rid]['rhash']) > 1 else _('아니오')

    return [v for k, v in peers.items()]


def get_all_info():
    devices = RustDesDevice.objects.all()
    peers = RustDeskPeer.objects.all()
    devices = {x.rid: model_to_dict2(x) for x in devices}
    now = datetime.datetime.now()
    for peer in peers:
        user = UserProfile.objects.filter(Q(id=peer.uid)).first()
        device = devices.get(peer.rid, None)
        if device:
            devices[peer.rid]['rust_user'] = user.username

    for rid in devices.keys():
        if not devices[rid].get('rust_user', ''):
            devices[rid]['rust_user'] = _('로그인 안됨')
    
    # 각 디바이스의 최근 상담원 번호 조회 (10분 이내만)
    ten_minutes_ago = now - datetime.timedelta(minutes=10)
    for rid in devices.keys():
        latest_agent = AgentConnectionLog.objects.filter(
            mdesk_id=rid,
            create_time__gte=ten_minutes_ago
        ).order_by('-create_time').first()
        devices[rid]['latest_agent_num'] = latest_agent.agent_num if latest_agent else '-'
    
    for k, v in devices.items():
        devices[k]['status'] = _('온라인') if (now - datetime.datetime.strptime(v['update_time'], '%Y-%m-%d %H:%M')).seconds <= 120 else _('오프라인')
    return [v for k, v in devices.items()]


@login_required(login_url='/api/user_action?action=login')
def work(request):
    username = request.user
    u = UserProfile.objects.get(username=username)

    show_type = request.GET.get('show_type', '')
    show_all = True if show_type == 'admin' and u.is_admin else False
    paginator = Paginator(get_all_info(), 15) if show_type == 'admin' and u.is_admin else Paginator(get_single_info(u.id), 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 최근 접속 원격지 및 연결 통계 (일반 사용자 뷰에서만)
    recent_remotes = get_recent_remotes_for_user(u.id) if not show_all else []
    connection_stats = get_connection_stats_for_user(u.id) if not show_all else {'today': 0, 'week': 0, 'total': 0}

    return render(request, 'show_work.html', {
        'u': u, 'show_all': show_all, 'page_obj': page_obj,
        'recent_remotes': recent_remotes,
        'connection_stats': connection_stats
    })


@login_required(login_url='/api/user_action?action=login')
def down_peers(request):
    username = request.user
    u = UserProfile.objects.get(username=username)

    if not u.is_admin:
        print(u.is_admin)
        return HttpResponseRedirect('/api/work')

    all_info = get_all_info()
    f = xlwt.Workbook(encoding='utf-8')
    sheet1 = f.add_sheet(_(u'기기 정보표'), cell_overwrite_ok=True)
    all_fields = [x.name for x in RustDesDevice._meta.get_fields()]
    all_fields.append('rust_user')
    for i, one in enumerate(all_info):
        for j, name in enumerate(all_fields):
            if i == 0:
                # 写入列名
                sheet1.write(i, j, name)
            sheet1.write(i + 1, j, one.get(name, '-'))

    sio = BytesIO()
    f.save(sio)
    sio.seek(0)
    response = HttpResponse(sio.getvalue(), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=DeviceInfo.xls'
    response.write(sio.getvalue())
    return response


def check_sharelink_expired(sharelink):
    now = datetime.datetime.now()
    if sharelink.create_time > now:
        return False
    if (now - sharelink.create_time).seconds < 15 * 60:
        return False
    else:
        sharelink.is_expired = True
        sharelink.save()
        return True


@login_required(login_url='/api/user_action?action=login')
def share(request):
    peers = RustDeskPeer.objects.filter(Q(uid=request.user.id))
    sharelinks = ShareLink.objects.filter(Q(uid=request.user.id) & Q(is_used=False) & Q(is_expired=False))

    # 省资源：处理已过期请求，不主动定时任务轮询请求，在任意地方请求时，检查是否过期，过期则保存。
    # now = datetime.datetime.now()
    for sl in sharelinks:
        check_sharelink_expired(sl)
    sharelinks = ShareLink.objects.filter(Q(uid=request.user.id) & Q(is_used=False) & Q(is_expired=False))
    peers = [{'id': ix + 1, 'name': f'{p.rid}|{p.alias}'} for ix, p in enumerate(peers)]
    sharelinks = [{'shash': s.shash, 'is_used': s.is_used, 'is_expired': s.is_expired, 'create_time': s.create_time, 'peers': s.peers} for ix, s in enumerate(sharelinks)]

    if request.method == 'GET':
        url = request.build_absolute_uri()
        if url.endswith('share'):
            return render(request, 'share.html', {'peers': peers, 'sharelinks': sharelinks})
        else:
            shash = url.split('/')[-1]
            sharelink = ShareLink.objects.filter(Q(shash=shash))
            msg = ''
            title = '성공'
            if not sharelink:
                title = '오류'
                msg = f'링크 {url}:<br>공유 링크가 존재하지 않거나 만료되었습니다.'
            else:
                sharelink = sharelink[0]
                if str(request.user.id) == str(sharelink.uid):
                    title = '오류'
                    msg = f'링크 {url}:<br><br>자신에게 링크를 공유할 수 없습니다!'
                else:
                    sharelink.is_used = True
                    sharelink.save()
                    peers = sharelink.peers
                    peers = peers.split(',')
                    # 自己的peers若重叠，需要跳过
                    peers_self_ids = [x.rid for x in RustDeskPeer.objects.filter(Q(uid=request.user.id))]
                    peers_share = RustDeskPeer.objects.filter(Q(rid__in=peers) & Q(uid=sharelink.uid))
                    # peers_share_ids = [x.rid for x in peers_share]

                    for peer in peers_share:
                        if peer.rid in peers_self_ids:
                            continue
                        # peer = RustDeskPeer.objects.get(rid=peer.rid)
                        peer_f = RustDeskPeer.objects.filter(Q(rid=peer.rid) & Q(uid=sharelink.uid))
                        if not peer_f:
                            msg += f"{peer.rid}이(가) 이미 존재합니다,"
                            continue

                        if len(peer_f) > 1:
                            msg += f'{peer.rid}이(가) 여러 개 존재합니다. 건너뜁니다. '
                            continue
                        peer = peer_f[0]
                        peer.id = None
                        peer.uid = request.user.id
                        peer.save()
                        msg += f"{peer.rid},"

                    msg += '성공적으로 가져왔습니다.'

            title = _(title)
            msg = _(msg)
            return render(request, 'msg.html', {'title': msg, 'msg': msg})
    else:
        data = request.POST.get('data', '[]')

        data = json.loads(data)
        if not data:
            return JsonResponse({'code': 0, 'msg': _('데이터가 비어있습니다.')})
        rustdesk_ids = [x['title'].split('|')[0] for x in data]
        rustdesk_ids = ','.join(rustdesk_ids)
        sharelink = ShareLink(
            uid=request.user.id,
            shash=getStrMd5(str(time.time()) + salt),
            peers=rustdesk_ids,
        )
        sharelink.save()

        return JsonResponse({'code': 1, 'shash': sharelink.shash})


def get_conn_log():
    logs = ConnLog.objects.all()
    logs = {x.id: model_to_dict(x) for x in logs}

    for k, v in logs.items():
        try:
            peer = RustDeskPeer.objects.get(rid=v['rid'])
            logs[k]['alias'] = peer.alias
        except: # noqa
            logs[k]['alias'] = _('UNKNOWN')
        try:
            peer = RustDeskPeer.objects.get(rid=v['from_id'])
            logs[k]['from_alias'] = peer.alias
        except: # noqa
            logs[k]['from_alias'] = _('UNKNOWN')
        # from_zone = tz.tzutc()
        # to_zone = tz.tzlocal()
        # utc = logs[k]['logged_at']
        # utc = utc.replace(tzinfo=from_zone)
        # logs[k]['logged_at'] = utc.astimezone(to_zone)
        try:
            duration = round((logs[k]['conn_end'] - logs[k]['conn_start']).total_seconds())
            m, s = divmod(duration, 60)
            h, m = divmod(m, 60)
            # d, h = divmod(h, 24)
            logs[k]['duration'] = f'{h:02d}:{m:02d}:{s:02d}'
        except:   # noqa
            logs[k]['duration'] = -1

    sorted_logs = sorted(logs.items(), key=lambda x: x[1]['conn_start'], reverse=True)
    new_ordered_dict = {}
    for key, alog in sorted_logs:
        new_ordered_dict[key] = alog

    return [v for k, v in new_ordered_dict.items()]


def get_file_log():
    logs = FileLog.objects.all()
    logs = {x.id: model_to_dict(x) for x in logs}

    for k, v in logs.items():
        try:
            peer_remote = RustDeskPeer.objects.get(rid=v['remote_id'])
            logs[k]['remote_alias'] = peer_remote.alias
        except:   # noqa
            logs[k]['remote_alias'] = _('UNKNOWN')
        try:
            peer_user = RustDeskPeer.objects.get(rid=v['user_id'])
            logs[k]['user_alias'] = peer_user.alias
        except:   # noqa
            logs[k]['user_alias'] = _('UNKNOWN')

    sorted_logs = sorted(logs.items(), key=lambda x: x[1]['logged_at'], reverse=True)
    new_ordered_dict = {}
    for key, alog in sorted_logs:
        new_ordered_dict[key] = alog

    return [v for k, v in new_ordered_dict.items()]


@login_required(login_url='/api/user_action?action=login')
def conn_log(request):
    paginator = Paginator(get_conn_log(), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'show_conn_log.html', {'page_obj': page_obj})


@login_required(login_url='/api/user_action?action=login')
def file_log(request):
    paginator = Paginator(get_file_log(), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'show_file_log.html', {'page_obj': page_obj})


def encrypt_password(password, key="rustdesk"):
    """암호 암호화 함수"""
    if not password:
        return ''
    # XOR 연산
    xor_bytes = bytes([ord(c) ^ ord(key[i % len(key)]) for i, c in enumerate(password)])
    # Base64 인코딩 (URL Safe, No Padding)
    return base64.urlsafe_b64encode(xor_bytes).decode().replace('=', '')


def custom_app(request):
    """커스텀 앱 설정 페이지 및 다운로드"""
    # POST 요청: 설정 저장 또는 실행파일 다운로드
    if request.method == 'POST':
        action = request.POST.get('action', '')
        
        if action == 'save':
            # 저장 요청은 admin.787.kr 도메인에서만 가능
            host = request.get_host()
            admin_host = getattr(settings, 'ID_SERVER', 'admin.787.kr')
            if not (host.startswith(admin_host) or host.startswith('admin.localhost')):
                return JsonResponse({'code': 0, 'msg': _('관리자 도메인에서만 접근 가능합니다.')})
            
            # 저장 요청은 로그인 필수
            if not request.user.is_authenticated:
                return JsonResponse({'code': 0, 'msg': _('로그인이 필요합니다.')})
            
            username = request.user.username
            u = UserProfile.objects.get(username=username)
            
            try:
                app_name = request.POST.get('app_name', 'MDesk')
                password = request.POST.get('password', '')
                title = request.POST.get('title', 'Your Desktop')
                description = request.POST.get('description', 'Your desktop can be accessed with this ID and password.')
                phone = request.POST.get('phone', '')
                logo_file = request.FILES.get('logo', None)
                notify_team = request.POST.get('notify_team', '0') == '1'
                
                # 사용자별로 하나만 생성/업데이트 (unique_together로 보장됨)
                custom_config, created = CustomAppConfig.objects.get_or_create(
                    uid=u,
                    defaults={'url_nickname': None}
                )
                
                # 암호 변경 여부 확인
                password_changed = custom_config.password != password
                
                # 설정 업데이트
                custom_config.app_name = app_name
                custom_config.password = password
                custom_config.title = title
                custom_config.description = description
                custom_config.phone = phone
                
                # 로고 파일이 업로드된 경우
                if logo_file:
                    # 기존 로고 삭제
                    if custom_config.logo:
                        try:
                            if os.path.exists(custom_config.logo.path):
                                os.remove(custom_config.logo.path)
                        except Exception as e:
                            print(f'기존 로고 삭제 오류: {str(e)}')
                    # upload_to 함수가 자동으로 uid를 파일명 앞에 추가함
                    custom_config.logo = logo_file
                
                # 데이터베이스에 저장
                custom_config.save()
                
                # 암호 변경 시 팀원에게 알림톡 발송
                notified_count = 0
                if password_changed and notify_team:
                    from api.models import TeamMember
                    team_members = TeamMember.objects.filter(user=u, is_active=True)
                    
                    for member in team_members:
                        if member.phone:
                            try:
                                # 카카오 알림톡 발송 (템플릿: MDTL05)
                                send_custom_app_password_notification(u.username, member.name, member.phone, password)
                                notified_count += 1
                            except Exception as e:
                                print(f"팀원 알림톡 발송 실패: {member.name} - {str(e)}")
                
                return JsonResponse({
                    'code': 1, 
                    'msg': _('설정이 성공적으로 저장되었습니다.'),
                    'created': created,
                    'notified_count': notified_count
                })
            except Exception as e:
                return JsonResponse({
                    'code': 0, 
                    'msg': _('저장 중 오류가 발생했습니다: {}').format(str(e))
                })
        
        elif action == 'download':
            # 실행파일 다운로드 (공개 지원 페이지에서도 호출 가능)
            target_username = request.POST.get('target_username', '')
            if target_username:
                u = UserProfile.objects.filter(username=target_username).first()
            elif request.user.is_authenticated:
                u = UserProfile.objects.get(username=request.user.username)
            else:
                return JsonResponse({'code': 0, 'msg': _('사용자 정보가 필요합니다.')})

            if not u:
                return JsonResponse({'code': 0, 'msg': _('사용자를 찾을 수 없습니다.')})
            
            # 실행파일 디렉토리 확인 및 생성
            executable_dir = getattr(settings, 'EXECUTABLE_DIR', os.path.join(settings.BASE_DIR, 'executables'))
            os.makedirs(executable_dir, exist_ok=True)
            
            # 템플릿 파일명 (기본값: rustdesk_portable-id=admin.exe)
            template_filename = getattr(settings, 'EXECUTABLE_TEMPLATE', 'rustdesk_portable-id={username}.exe')
            template_path = os.path.join(executable_dir, template_filename.format(username='admin'))
            
            # 템플릿 파일이 없으면 기본 파일명으로 시도
            if not os.path.exists(template_path):
                # 다른 가능한 파일명들 시도
                possible_names = [
                    'MDesk_portable-id=admin.exe',
                    'MDesk_portable.exe',
                    'MDesk.exe'
                ]
                template_path = None
                for name in possible_names:
                    test_path = os.path.join(executable_dir, name)
                    if os.path.exists(test_path):
                        template_path = test_path
                        break
                
                if not template_path:
                    return JsonResponse({
                        'code': 0, 
                        'msg': _('실행파일을 찾을 수 없습니다. 관리자에게 문의하세요.'),
                        'file_location': executable_dir
                    })
            
            # 사용자별 파일명 생성
            username = u.username
            download_filename = f'MDesk_portable-id={username}.exe'
            
            try:
                # 실행파일 읽기
                with open(template_path, 'rb') as f:
                    file_content = f.read()
                
                # 다운로드 응답 생성
                response = HttpResponse(file_content, content_type='application/x-msdownload')
                response['Content-Disposition'] = f'attachment; filename="{download_filename}"'
                response['Content-Length'] = len(file_content)
                
                return response
            except Exception as e:
                return JsonResponse({
                    'code': 0, 
                    'msg': _('실행파일 다운로드 중 오류가 발생했습니다: {}').format(str(e)),
                    'file_location': executable_dir
                })
        
        return JsonResponse({'code': 0, 'msg': _('잘못된 요청입니다.')})
    
    # GET 요청: 폼 표시 (로그인 필수)
    if not request.user.is_authenticated:
        return redirect('/api/user_action?action=login')
    
    try:
        u = UserProfile.objects.get(username=request.user.username)
    except UserProfile.DoesNotExist:
        print(f'[custom_app 오류] 사용자를 찾을 수 없습니다: {request.user.username}')
        return HttpResponse(_('사용자 정보를 찾을 수 없습니다.'), status=404)
    except Exception as e:
        print(f'[custom_app 오류] 사용자 조회 실패: {str(e)}')
        return HttpResponse(_('오류가 발생했습니다: {}').format(str(e)), status=500)
    
    try:
        # url_nickname 중복 문제를 방지하기 위해 defaults 지정
        custom_config, created = CustomAppConfig.objects.get_or_create(
            uid=u,
            defaults={'url_nickname': None}
        )
    except Exception as e:
        print(f'[custom_app 오류] CustomAppConfig 조회/생성 실패: {str(e)}')
        import traceback
        print(traceback.format_exc())
        return HttpResponse(_('설정을 불러오는 중 오류가 발생했습니다: {}').format(str(e)), status=500)
    
    try:
        # 암호화된 암호 계산 (미리보기용)
        encrypted_password = encrypt_password(custom_config.password) if custom_config.password else '8bht8f'
        
        # 로고 URL 안전하게 처리
        logo_url = ''
        if custom_config.logo:
            try:
                if custom_config.logo.name and os.path.exists(custom_config.logo.path):
                    logo_url = custom_config.logo.url
                else:
                    # 파일이 없으면 로고 필드 초기화
                    custom_config.logo = None
                    custom_config.save(update_fields=['logo'])
            except Exception as logo_error:
                print(f'[custom_app 경고] 로고 파일 확인 실패: {str(logo_error)}')
                # 파일이 없으면 로고 필드 초기화
                try:
                    custom_config.logo = None
                    custom_config.save(update_fields=['logo'])
                except:
                    pass
        
        # 팀원 목록 조회
        from api.models import TeamMember
        team_members = TeamMember.objects.filter(user=u, is_active=True)
        
        return render(request, 'custom_app.html', {
            'u': u, 
            'config': custom_config, 
            'encrypted_password': encrypted_password,
            'logo_url': logo_url,
            'team_members': team_members
        })
    except Exception as e:
        print(f'[custom_app 오류] 템플릿 렌더링 실패: {str(e)}')
        import traceback
        print(traceback.format_exc())
        return HttpResponse(_('페이지를 불러오는 중 오류가 발생했습니다: {}').format(str(e)), status=500)


def default_page(request):
    """기본 다운로드 페이지"""
    return render(request, 'default.html')


def public_support(request, username):
    """공개 지원 페이지 (상담원 정보 표시)
    
    username 또는 user_pkid(숫자)로 접근 가능
    예: /imedix 또는 /3 (숫자로 접근 시 username으로 리다이렉트)
    """
    # 숫자인 경우 user_pkid로 검색 후 username으로 리다이렉트
    if username.isdigit():
        user = UserProfile.objects.filter(id=int(username)).first()
        if user:
            return redirect(f'/{user.username}')
        else:
            return HttpResponse(_("사용자를 찾을 수 없습니다."), status=404)
    
    # 문자열인 경우 username으로 검색
    user = UserProfile.objects.filter(username=username).first()
    if not user:
        return HttpResponse(_("사용자를 찾을 수 없습니다."), status=404)
    
    # 커스텀 설정 가져오기 (없으면 기본값)
    custom_config = CustomAppConfig.objects.filter(uid=user).first()
    
    # 상담원 목록 가져오기
    agents = SupportAgent.objects.filter(uid=user)
    
    return render(request, 'support_page.html', {
        'target_user': user,
        'config': custom_config,
        'agents': agents,
    })


def download_default(request):
    """
    기본 MDesk 실행파일 다운로드 (로그인 불필요)
    /api/download_default 또는 default 페이지에서 사용
    """
    # 실행파일 디렉토리 확인
    executable_dir = getattr(settings, 'EXECUTABLE_DIR', os.path.join(settings.BASE_DIR, 'executables'))
    
    # 가능한 파일명들 시도
    possible_names = [
        'MDesk_portable.exe',
        'MDesk_portable-id=admin.exe',
        'MDesk.exe',
        'rustdesk_portable.exe'
    ]
    
    template_path = None
    for name in possible_names:
        path = os.path.join(executable_dir, name)
        if os.path.exists(path):
            template_path = path
            break
    
    if not template_path:
        return HttpResponse(_('실행파일을 찾을 수 없습니다. 관리자에게 문의하세요.'), status=404)
    
    try:
        # 실행파일 읽기
        with open(template_path, 'rb') as f:
            file_content = f.read()
        
        # 다운로드 응답 생성
        response = HttpResponse(file_content, content_type='application/x-msdownload')
        response['Content-Disposition'] = 'attachment; filename="MDesk_portable.exe"'
        response['Content-Length'] = len(file_content)
        
        return response
    except Exception as e:
        return HttpResponse(_('파일 다운로드 중 오류가 발생했습니다: {}').format(str(e)), status=500)


@login_required(login_url='/api/user_action?action=login')
def group_manage(request):
    """그룹관리 페이지"""
    username = request.user
    u = UserProfile.objects.get(username=username)
    
    # 최고관리자만 접근 가능
    if not u.is_superuser:
        return HttpResponseRedirect('/api/work')
    
    # 그룹 목록 조회 (삭제되지 않은 것만)
    groups = Group.objects.filter(uid=u, is_deleted=False).order_by('-create_time')
    
    # 수정 모드인 경우
    edit_id = request.GET.get('edit', None)
    edit_group = None
    if edit_id:
        edit_group = Group.objects.filter(id=edit_id, uid=u, is_deleted=False).first()
    
    return render(request, 'group_manage.html', {
        'u': u,
        'groups': groups,
        'edit_group': edit_group
    })


@login_required(login_url='/api/user_action?action=login')
def group_save(request):
    """그룹 저장 API"""
    result = {'code': 0, 'msg': ''}
    
    if request.method != 'POST':
        result['msg'] = _('잘못된 요청입니다.')
        return JsonResponse(result)
    
    username = request.user
    u = UserProfile.objects.get(username=username)
    
    # 최고관리자만 접근 가능
    if not u.is_superuser:
        result['msg'] = _('권한이 없습니다.')
        return JsonResponse(result)
    
    group_id = request.POST.get('group_id', '')
    company_name = request.POST.get('company_name', '')
    business_number = request.POST.get('business_number', '')
    representative = request.POST.get('representative', '')
    contact_name = request.POST.get('contact_name', '')
    contact_phone = request.POST.get('contact_phone', '')
    contact_email = request.POST.get('contact_email', '')
    address = request.POST.get('address', '')
    address_detail = request.POST.get('address_detail', '')
    memo = request.POST.get('memo', '')
    
    if not company_name:
        result['msg'] = _('회사명을 입력해주세요.')
        return JsonResponse(result)
    
    try:
        if group_id:
            # 수정
            group = Group.objects.filter(id=group_id, uid=u, is_deleted=False).first()
            if not group:
                result['msg'] = _('그룹을 찾을 수 없습니다.')
                return JsonResponse(result)
            
            group.company_name = company_name
            group.business_number = business_number
            group.representative = representative
            group.contact_name = contact_name
            group.contact_phone = contact_phone
            group.contact_email = contact_email
            group.address = address
            group.address_detail = address_detail
            group.memo = memo
            group.save()
            
            result['code'] = 1
            result['msg'] = _('그룹이 수정되었습니다.')
        else:
            # 신규 등록
            group = Group(
                uid=u,
                company_name=company_name,
                business_number=business_number,
                representative=representative,
                contact_name=contact_name,
                contact_phone=contact_phone,
                contact_email=contact_email,
                address=address,
                address_detail=address_detail,
                memo=memo
            )
            group.save()
            
            result['code'] = 1
            result['msg'] = _('그룹이 등록되었습니다.')
            result['group_id'] = group.id
    except Exception as e:
        result['msg'] = _('저장 중 오류가 발생했습니다: {}').format(str(e))
    
    return JsonResponse(result)


@login_required(login_url='/api/user_action?action=login')
def group_delete(request, group_id):
    """그룹 소프트 삭제 API"""
    result = {'code': 0, 'msg': ''}
    
    username = request.user
    u = UserProfile.objects.get(username=username)
    
    # 최고관리자만 접근 가능
    if not u.is_superuser:
        result['msg'] = _('권한이 없습니다.')
        return JsonResponse(result)
    
    try:
        group = Group.objects.filter(id=group_id, uid=u, is_deleted=False).first()
        if not group:
            result['msg'] = _('그룹을 찾을 수 없습니다.')
            return JsonResponse(result)
        
        # 소프트 삭제 (실제 삭제 대신 플래그 변경)
        group.soft_delete()
        result['code'] = 1
        result['msg'] = _('그룹이 삭제되었습니다.')
    except Exception as e:
        result['msg'] = _('삭제 중 오류가 발생했습니다: {}').format(str(e))
    
    return JsonResponse(result)


@login_required(login_url='/api/user_action?action=login')
def user_manage(request):
    """사용자관리 페이지"""
    username = request.user
    u = UserProfile.objects.get(username=username)
    
    # 최고관리자 또는 그룹관리자만 접근 가능
    if not (u.is_superuser or u.is_group_admin):
        return HttpResponseRedirect('/api/work')
    
    # 사용자 목록 조회
    if u.is_superuser:
        # 최고관리자는 모든 사용자 조회 (자신 제외)
        users = UserProfile.objects.exclude(id=u.id).order_by('-is_admin', '-is_active', 'username')
    else:
        # 그룹관리자는 자신의 소속 유저만 조회
        # 같은 group_id 또는 같은 company_name을 가진 유저들
        users_query = UserProfile.objects.exclude(id=u.id)
        
        # group_id가 있는 경우 같은 group_id를 가진 유저들
        if u.group_id:
            users_query = users_query.filter(group_id=u.group_id)
        # company_name이 있는 경우 같은 company_name을 가진 유저들
        elif u.company_name:
            users_query = users_query.filter(company_name=u.company_name)
        else:
            # group_id와 company_name이 모두 없는 경우 빈 목록
            users_query = users_query.none()
        
        users = users_query.order_by('-is_admin', '-is_active', 'username')
    
    # 그룹(회사) 목록 조회 (콤보박스용)
    groups = Group.objects.filter(uid=u, is_deleted=False).order_by('company_name')
    
    # 릴레이 서버 목록 조회
    relay_servers = RelayServer.objects.filter(is_active=True).order_by('server_address')
    
    # 수정 모드인 경우
    edit_id = request.GET.get('edit', None)
    edit_user = None
    if edit_id:
        edit_user = UserProfile.objects.filter(id=edit_id).exclude(id=u.id).first()
        # 그룹관리자인 경우 자신의 소속 유저만 수정 가능
        if u.is_group_admin and not u.is_superuser:
            if edit_user:
                if u.group_id and edit_user.group_id != u.group_id:
                    edit_user = None
                elif u.company_name and edit_user.company_name != u.company_name:
                    edit_user = None
    
    return render(request, 'user_manage.html', {
        'u': u,
        'users': users,
        'edit_user': edit_user,
        'groups': groups,
        'relay_servers': relay_servers
    })


@login_required(login_url='/api/user_action?action=login')
def user_manage_save(request):
    """사용자 저장 API"""
    result = {'code': 0, 'msg': ''}
    
    if request.method != 'POST':
        result['msg'] = _('잘못된 요청입니다.')
        return JsonResponse(result)
    
    username = request.user
    u = UserProfile.objects.get(username=username)
    
    # 최고관리자 또는 그룹관리자만 접근 가능
    if not (u.is_superuser or u.is_group_admin):
        result['msg'] = _('권한이 없습니다.')
        return JsonResponse(result)
    
    user_id = request.POST.get('user_id', '')
    target_username = request.POST.get('username', '')
    company_name = request.POST.get('company_name', '')
    group_id = request.POST.get('group_id', '')
    email = request.POST.get('email', '')
    phone = request.POST.get('phone', '')
    membership_level = request.POST.get('membership_level', 'free')
    max_agents = request.POST.get('max_agents', '3')
    
    # 릴레이 서버 기본값 처리
    default_server = RelayServer.objects.filter(is_default=True).first()
    default_address = default_server.server_address if default_server else '222.239.231.91'
    # 그룹관리자는 릴레이 서버 변경 불가 (기본값 사용)
    if u.is_group_admin and not u.is_superuser:
        relay_server = default_address
    else:
        relay_server = request.POST.get('relay_server', default_address)
    
    is_active = request.POST.get('is_active', '0') == '1'
    is_admin = request.POST.get('is_admin', '0') == '1'
    is_group_admin = request.POST.get('is_group_admin', '0') == '1'
    new_password = request.POST.get('new_password', '')
    email_2fa = request.POST.get('email_2fa', '0') == '1' and bool(email)
    phone_2fa = request.POST.get('phone_2fa', '0') == '1' and bool(phone)
    if email_2fa and phone_2fa:
        email_2fa = False
    
    try:
        max_agents = int(max_agents)
    except:
        max_agents = 3
    
    # group_id 처리
    try:
        group_id = int(group_id) if group_id else None
    except:
        group_id = None
    
    try:
        if user_id:
            # 수정
            target_user = UserProfile.objects.filter(id=user_id).exclude(id=u.id).first()
            if not target_user:
                result['msg'] = _('사용자를 찾을 수 없습니다.')
                return JsonResponse(result)
            
            # 그룹관리자인 경우 자신의 소속 유저만 수정 가능
            if u.is_group_admin and not u.is_superuser:
                if u.group_id and target_user.group_id != u.group_id:
                    result['msg'] = _('권한이 없습니다. 자신의 소속 유저만 수정할 수 있습니다.')
                    return JsonResponse(result)
                elif u.company_name and target_user.company_name != u.company_name:
                    result['msg'] = _('권한이 없습니다. 자신의 소속 유저만 수정할 수 있습니다.')
                    return JsonResponse(result)
                # 그룹관리자는 회사명/그룹 변경 불가 (자신의 그룹으로 고정)
                company_name = u.company_name if u.company_name else company_name
                group_id = u.group_id if u.group_id else group_id
            
            # 최고관리자는 이메일/전화번호 중복 체크 없이 저장 가능
            target_user.company_name = company_name
            target_user.group_id = group_id
            target_user.email = email
            target_user.phone = phone
            target_user.email_2fa = email_2fa
            target_user.phone_2fa = phone_2fa
            # 그룹관리자는 회원등급 변경 불가 (기존 값 유지)
            if not (u.is_group_admin and not u.is_superuser):
                target_user.membership_level = membership_level
            # 그룹관리자는 최대 상담원 변경 불가 (기존 값 유지)
            if not (u.is_group_admin and not u.is_superuser):
                target_user.max_agents = max_agents
            # 그룹관리자는 릴레이 서버 변경 불가 (기존 값 유지)
            if not (u.is_group_admin and not u.is_superuser):
                target_user.relay_server = relay_server
            target_user.is_active = is_active
            
            # 관리자 권한 설정 (최고관리자만 가능)
            if u.is_superuser:
                target_user.is_admin = is_admin
                # is_group_admin 설정 (is_admin = true인 유저만 설정 가능)
                if u.is_admin:
                    target_user.is_group_admin = is_group_admin
            else:
                # 그룹관리자는 관리자 권한 변경 불가 (기존 값 유지)
                pass
            
            # 비밀번호 변경
            password_changed = False
            if new_password:
                if len(new_password) < 8 or len(new_password) > 20:
                    result['msg'] = _('비밀번호 길이가 요구사항에 맞지 않습니다. 8~20자여야 합니다.')
                    return JsonResponse(result)
                target_user.set_password(new_password)
                password_changed = True
            
            target_user.save()
            
            # 비밀번호 변경 시 기존 로그인 토큰 무효화 (강제 로그아웃)
            if password_changed:
                from api.models import RustDeskToken
                RustDeskToken.objects.filter(uid=target_user.id).delete()
            
            result['code'] = 1
            result['msg'] = _('사용자 정보가 수정되었습니다.')
        else:
            # 신규 등록
            if not target_username:
                result['msg'] = _('사용자명을 입력해주세요.')
                return JsonResponse(result)
            
            if len(target_username) <= 3:
                result['msg'] = _('사용자명은 3자 이상이어야 합니다.')
                return JsonResponse(result)
            
            # 사용자명 중복 확인
            if UserProfile.objects.filter(username=target_username).exists():
                result['msg'] = _('이미 존재하는 사용자명입니다.')
                return JsonResponse(result)
            
            if not new_password:
                result['msg'] = _('비밀번호를 입력해주세요.')
                return JsonResponse(result)
            
            if len(new_password) < 8 or len(new_password) > 20:
                result['msg'] = _('비밀번호 길이가 요구사항에 맞지 않습니다. 8~20자여야 합니다.')
                return JsonResponse(result)
            
            # 그룹관리자인 경우 자신의 소속으로만 등록 가능
            if u.is_group_admin and not u.is_superuser:
                # group_id가 있는 경우 같은 group_id로만 등록 가능
                if u.group_id:
                    # 그룹관리자는 회사명/그룹 변경 불가 (자신의 그룹으로 고정)
                    company_name = u.company_name if u.company_name else company_name
                    group_id = u.group_id
                # company_name이 있는 경우 같은 company_name으로만 등록 가능
                elif u.company_name:
                    # 그룹관리자는 회사명 변경 불가 (자신의 회사명으로 고정)
                    company_name = u.company_name
                    group_id = None
                # group_id와 company_name이 모두 없는 경우 등록 불가
                else:
                    result['msg'] = _('권한이 없습니다. 소속 정보가 없어 등록할 수 없습니다.')
                    return JsonResponse(result)
            
            # 최고관리자는 이메일/전화번호 중복 체크 없이 저장 가능
            # 그룹관리자는 회원등급을 기본값(free)으로 고정
            final_membership_level = 'free' if (u.is_group_admin and not u.is_superuser) else membership_level
            # 그룹관리자는 최대 상담원을 기본값(3)으로 고정
            final_max_agents = 3 if (u.is_group_admin and not u.is_superuser) else max_agents
            
            new_user = UserProfile(
                username=target_username,
                company_name=company_name,
                group_id=group_id,
                email=email,
                phone=phone,
                email_2fa=email_2fa,
                phone_2fa=phone_2fa,
                membership_level=final_membership_level,
                max_agents=final_max_agents,
                relay_server=relay_server,
                is_active=is_active,
                is_admin=is_admin if u.is_superuser else False,  # 그룹관리자는 관리자 권한 부여 불가
                is_group_admin=False  # 그룹관리자는 is_group_admin 설정 불가
            )
            new_user.set_password(new_password)
            new_user.save()
            
            result['code'] = 1
            result['msg'] = _('사용자가 등록되었습니다.')
            result['user_id'] = new_user.id
    except Exception as e:
        result['msg'] = _('저장 중 오류가 발생했습니다: {}').format(str(e))
    
    return JsonResponse(result)


@login_required(login_url='/api/user_action?action=login')
def user_manage_delete(request, user_id):
    """사용자 비활성화 API"""
    result = {'code': 0, 'msg': ''}
    
    username = request.user
    u = UserProfile.objects.get(username=username)
    
    # 최고관리자만 접근 가능
    if not u.is_superuser:
        result['msg'] = _('권한이 없습니다.')
        return JsonResponse(result)
    
    try:
        target_user = UserProfile.objects.filter(id=user_id).exclude(id=u.id).first()
        if not target_user:
            result['msg'] = _('사용자를 찾을 수 없습니다.')
            return JsonResponse(result)
        
        # superuser는 삭제 불가
        if target_user.is_superuser:
            result['msg'] = _('최고관리자는 삭제할 수 없습니다.')
            return JsonResponse(result)
        
        # 비활성화 처리 (소프트 삭제)
        target_user.is_active = False
        target_user.save()
        
        result['code'] = 1
        result['msg'] = _('사용자가 비활성화되었습니다.')
    except Exception as e:
        result['msg'] = _('삭제 중 오류가 발생했습니다: {}').format(str(e))
    
    return JsonResponse(result)


@login_required(login_url='/api/user_action?action=login')
def relay_server_list(request):
    """릴레이 서버 목록 조회 API"""
    servers = RelayServer.objects.all().order_by('-is_default', '-create_time')
    data = []
    for s in servers:
        data.append({
            'id': s.id,
            'server_address': s.server_address,
            'server_name': s.server_name,
            'public_key': s.public_key,
            'is_default': s.is_default,
            'is_active': s.is_active
        })
    return JsonResponse({'code': 0, 'msg': '', 'count': len(data), 'data': data})


@login_required(login_url='/api/user_action?action=login')
def relay_server_set_default(request, id):
    """릴레이 서버 기본값 설정 API"""
    try:
        server = RelayServer.objects.get(id=id)
        server.is_default = True
        server.save()  # 모델의 save 메서드에서 다른 서버의 is_default를 False로 만듭니다.
        return JsonResponse({'code': 1, 'msg': '기본 서버로 설정되었습니다.'})
    except RelayServer.DoesNotExist:
        return JsonResponse({'code': 0, 'msg': '존재하지 않는 서버입니다.'})


@login_required(login_url='/api/user_action?action=login')
def relay_server_add(request):
    """릴레이 서버 추가 API"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': 'POST 요청만 가능합니다.'})
    
    address = request.POST.get('address', '').strip()
    name = request.POST.get('name', '').strip()
    public_key = request.POST.get('public_key', '').strip()
    
    if not address:
        return JsonResponse({'code': 0, 'msg': '서버 주소를 입력해주세요.'})
    
    if RelayServer.objects.filter(server_address=address).exists():
        return JsonResponse({'code': 0, 'msg': '이미 존재하는 서버 주소입니다.'})
    
    # 첫 번째 서버이거나 요청이 있는 경우 기본값으로 설정
    is_first = not RelayServer.objects.exists()
    RelayServer.objects.create(server_address=address, server_name=name, public_key=public_key, is_default=is_first)
    return JsonResponse({'code': 1, 'msg': '릴레이 서버가 추가되었습니다.'})


@login_required(login_url='/api/user_action?action=login')
def relay_server_delete(request, id):
    """릴레이 서버 삭제 API"""
    try:
        server = RelayServer.objects.get(id=id)
        server.delete()
        return JsonResponse({'code': 1, 'msg': '삭제되었습니다.'})
    except RelayServer.DoesNotExist:
        return JsonResponse({'code': 0, 'msg': '존재하지 않는 서버입니다.'})


@login_required(login_url='/api/user_action?action=login')
def user_bulk_action(request):
    """사용자 일괄 처리 API"""
    result = {'code': 0, 'msg': ''}
    
    if request.method != 'POST':
        result['msg'] = _('잘못된 요청입니다.')
        return JsonResponse(result)
    
    username = request.user
    u = UserProfile.objects.get(username=username)
    
    # 최고관리자 또는 그룹관리자만 접근 가능
    if not (u.is_superuser or u.is_group_admin):
        result['msg'] = _('권한이 없습니다.')
        return JsonResponse(result)
    
    action = request.POST.get('action', '')
    user_ids_str = request.POST.get('user_ids', '[]')
    
    try:
        user_ids = json.loads(user_ids_str)
        if not isinstance(user_ids, list) or len(user_ids) == 0:
            result['msg'] = _('선택된 사용자가 없습니다.')
            return JsonResponse(result)
        
        # 최고관리자 제외
        target_users = UserProfile.objects.filter(id__in=user_ids, is_superuser=False)
        
        # 그룹관리자인 경우 자신의 소속 유저만 처리 가능
        if u.is_group_admin and not u.is_superuser:
            if u.group_id:
                target_users = target_users.filter(group_id=u.group_id)
            elif u.company_name:
                target_users = target_users.filter(company_name=u.company_name)
            else:
                result['msg'] = _('권한이 없습니다. 소속 정보가 없습니다.')
                return JsonResponse(result)
        
        if target_users.count() == 0:
            result['msg'] = _('처리할 수 있는 사용자가 없습니다.')
            return JsonResponse(result)
        
        if action == 'set_group':
            # 그룹 변경
            group_id = request.POST.get('group_id', '')
            if not group_id:
                result['msg'] = _('그룹을 선택해주세요.')
                return JsonResponse(result)
            
            try:
                # 최고관리자는 모든 그룹에 접근 가능
                group = Group.objects.get(id=group_id, is_deleted=False)
            except Group.DoesNotExist:
                result['msg'] = _('그룹을 찾을 수 없습니다.')
                return JsonResponse(result)
            
            updated_count = target_users.update(group_id=group.id, company_name=group.company_name)
            result['code'] = 1
            result['msg'] = f'{updated_count}명의 사용자가 "{group.company_name}" 그룹으로 변경되었습니다.'
            
        elif action == 'set_status':
            # 활성화/비활성화
            is_active = request.POST.get('is_active', '0') == '1'
            updated_count = target_users.update(is_active=is_active)
            status_text = '활성화' if is_active else '비활성화'
            result['code'] = 1
            result['msg'] = f'{updated_count}명의 사용자가 {status_text}되었습니다.'
            
        elif action == 'change_password':
            # 일괄 암호변경
            new_password = request.POST.get('new_password', '')
            notify_kakao = request.POST.get('notify_kakao', '0') == '1'
            
            if not new_password:
                result['msg'] = _('비밀번호를 입력해주세요.')
                return JsonResponse(result)

            if len(new_password) < 8 or len(new_password) > 20:
                result['msg'] = _('비밀번호는 8~20자여야 합니다.')
                return JsonResponse(result)

            # 각 사용자의 비밀번호 변경
            updated_count = 0
            kakao_sent_count = 0
            kakao_failed_users = []
            
            # 토큰 무효화를 위한 import
            from api.models import RustDeskToken
            
            for user in target_users:
                user.set_password(new_password)
                user.save()
                
                # 기존 로그인 토큰 무효화 (강제 로그아웃)
                RustDeskToken.objects.filter(uid=user.id).delete()
                
                updated_count += 1
                
                # 카카오톡 알림 발송
                if notify_kakao and user.phone:
                    try:
                        kakao_result = send_kakao_password_notification(
                            phone=user.phone,
                            username=user.username,
                            new_password=new_password,
                            company_name=user.company_name or ''
                        )
                        if kakao_result:
                            kakao_sent_count += 1
                        else:
                            kakao_failed_users.append(user.username)
                    except Exception as e:
                        kakao_failed_users.append(user.username)

            result['code'] = 1
            result['msg'] = f'{updated_count}명의 사용자 비밀번호가 변경되었습니다.'
            
            if notify_kakao:
                if kakao_sent_count > 0:
                    result['msg'] += f' (카카오톡 알림 {kakao_sent_count}건 발송)'
                if kakao_failed_users:
                    result['msg'] += f' (알림 실패: {", ".join(kakao_failed_users[:3])}{"..." if len(kakao_failed_users) > 3 else ""})'
            
        else:
            result['msg'] = _('알 수 없는 작업입니다.')
            
    except json.JSONDecodeError:
        result['msg'] = _('잘못된 데이터 형식입니다.')
    except Exception as e:
        result['msg'] = _('처리 중 오류가 발생했습니다: {}').format(str(e))
    
    return JsonResponse(result)


@login_required(login_url='/api/user_action?action=login')
def user_toggle_admin(request, user_id):
    """관리자 권한 토글 API"""
    result = {'code': 0, 'msg': ''}
    
    username = request.user
    u = UserProfile.objects.get(username=username)
    
    # 최고관리자만 접근 가능
    if not u.is_superuser:
        result['msg'] = _('권한이 없습니다.')
        return JsonResponse(result)
    
    try:
        target_user = UserProfile.objects.filter(id=user_id).exclude(id=u.id).first()
        if not target_user:
            result['msg'] = _('사용자를 찾을 수 없습니다.')
            return JsonResponse(result)
        
        target_user.is_admin = not target_user.is_admin
        target_user.save()
        
        status = _('관리자') if target_user.is_admin else _('일반 사용자')
        result['code'] = 1
        result['msg'] = _('{}님이 {}로 변경되었습니다.').format(target_user.username, status)
        result['is_admin'] = target_user.is_admin
    except Exception as e:
        result['msg'] = _('변경 중 오류가 발생했습니다: {}').format(str(e))
    
    return JsonResponse(result)


@login_required(login_url='/api/user_action?action=login')
def team_member_add(request):
    """팀원 추가 API"""
    result = {'code': 0, 'msg': ''}
    
    if request.method != 'POST':
        result['msg'] = _('잘못된 요청입니다.')
        return JsonResponse(result)
    
    username = request.user
    u = UserProfile.objects.get(username=username)
    
    name = request.POST.get('name', '').strip()
    phone = request.POST.get('phone', '').strip()
    memo = request.POST.get('memo', '').strip()
    
    if not name:
        result['msg'] = _('이름을 입력해주세요.')
        return JsonResponse(result)
    
    try:
        from api.models import TeamMember
        member = TeamMember.objects.create(
            user=u,
            name=name,
            phone=phone,
            memo=memo
        )
        result['code'] = 1
        result['msg'] = _('팀원이 추가되었습니다.')
        result['member_id'] = member.id
    except Exception as e:
        result['msg'] = _('추가 중 오류가 발생했습니다: {}').format(str(e))
    
    return JsonResponse(result)


@login_required(login_url='/api/user_action?action=login')
def team_member_delete(request, member_id):
    """팀원 삭제 API"""
    result = {'code': 0, 'msg': ''}
    
    username = request.user
    u = UserProfile.objects.get(username=username)
    
    try:
        from api.models import TeamMember
        member = TeamMember.objects.filter(id=member_id, user=u).first()
        if not member:
            result['msg'] = _('팀원을 찾을 수 없습니다.')
            return JsonResponse(result)
        
        member.delete()
        result['code'] = 1
        result['msg'] = _('팀원이 삭제되었습니다.')
    except Exception as e:
        result['msg'] = _('삭제 중 오류가 발생했습니다: {}').format(str(e))
    
    return JsonResponse(result)


def send_custom_app_password_notification(username, member_name, phone, new_password):
    """나만의 원격 암호 변경 알림톡 발송 (템플릿: MDTL05)"""
    import requests
    import json
    
    # 전화번호 정규화
    phone = phone.replace('-', '').replace(' ', '')
    
    # 카카오 알림톡 API 호출
    api_url = settings.KAKAO_ALIMTALK_API_URL
    if not api_url:
        return False
    
    # 카카오 알림톡 API 요청 데이터 구성
    data = {
        "titleinfo": {
            "id": settings.KAKAO_ALIMTALK_ID,
            "pw": settings.KAKAO_ALIMTALK_PW,
            "temNum": settings.KAKAO_ALIMTALK_TEM_NUM
        },
        "messageList": [
            {
                "rcvNum": phone,
                "title": new_password,
                "btnURL1": "",
                "msg": """"나만의 원격"앱의 공통 비밀번호가 변경되었습니다.
팀원분들은 서비스시 참고바랍니다.""",
                "ptno": "",
                "rcvDate": "",
                "rcvTime": "",
                "msgtype": "KKO",
                "sendNum": settings.KAKAO_ALIMTALK_SEND_NUM,
                "temCode": "MDTL05"
            }
        ]
    }
    
    try:
        # JSON 문자열로 변환 후 MSG 키에 담아서 POST 전송
        json_data = json.dumps(data, ensure_ascii=False)
        print(f"  - 요청 JSON 데이터: {json_data}")
        
        response = requests.post(
            api_url, 
            data={'MSG': json_data},
            headers={'Content-Type': 'application/x-www-form-urlencoded; charset=utf-8'},
            timeout=10
        )
        print(f"나만의 원격 알림톡 발송: {member_name}({phone}) - {response.text}")
        return True
    except Exception as e:
        print(f"나만의 원격 알림톡 발송 오류: {str(e)}")
        return False


def device_manage(request):
    """기기 관리 페이지"""
    u = request.user
    if not u.is_authenticated:
        return redirect('/api/user_action?action=login')
    
    # 관리자만 접근 가능
    try:
        profile = UserProfile.objects.get(username=u.username)
        if not profile.is_admin:
            return redirect('/api/work')
    except:
        return redirect('/api/work')
    
    # 검색/필터 파라미터
    search = request.GET.get('search', '')
    selected_status = request.GET.get('status', '')
    page = request.GET.get('page', 1)
    
    # 기기 목록 조회
    devices = RustDesDevice.objects.all().order_by('-update_time')
    
    # 검색 필터
    if search:
        devices = devices.filter(
            Q(rid__icontains=search) |
            Q(username__icontains=search) |
            Q(hostname__icontains=search) |
            Q(ip_address__icontains=search) |
            Q(os__icontains=search) |
            Q(version__icontains=search)
        )
    
    # 상태 필터 (최근 5분 이내 업데이트 = 온라인)
    online_threshold = datetime.datetime.now() - datetime.timedelta(minutes=5)
    
    if selected_status == 'online':
        devices = devices.filter(update_time__gte=online_threshold)
    elif selected_status == 'offline':
        devices = devices.filter(update_time__lt=online_threshold)
    
    # 통계
    all_devices = RustDesDevice.objects.all()
    total_count = all_devices.count()
    online_count = all_devices.filter(update_time__gte=online_threshold).count()
    offline_count = total_count - online_count
    
    # 온라인 상태 추가
    device_list = []
    for device in devices:
        device.is_online = device.update_time >= online_threshold if device.update_time else False
        device_list.append(device)
    
    # 페이징
    paginator = Paginator(device_list, 50)  # 페이지당 50개
    page_obj = paginator.get_page(page)
    
    return render(request, 'device_manage.html', {
        'u': profile,
        'devices': page_obj,
        'page_obj': page_obj,
        'total_count': total_count,
        'online_count': online_count,
        'offline_count': offline_count,
        'search': search,
        'selected_status': selected_status,
    })


def device_manage_delete(request, rid):
    """기기 삭제 API"""
    u = request.user
    if not u.is_authenticated:
        return JsonResponse({'code': 0, 'msg': '로그인이 필요합니다.'})
    
    try:
        profile = UserProfile.objects.get(username=u.username)
        if not profile.is_admin:
            return JsonResponse({'code': 0, 'msg': '권한이 없습니다.'})
    except:
        return JsonResponse({'code': 0, 'msg': '권한이 없습니다.'})
    
    try:
        device = RustDesDevice.objects.filter(rid=rid).first()
        if device:
            device.delete()
            return JsonResponse({'code': 1, 'msg': '삭제되었습니다.'})
        else:
            return JsonResponse({'code': 0, 'msg': '기기를 찾을 수 없습니다.'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})


def token_manage(request):
    """토큰 관리 페이지"""
    u = request.user
    if not u.is_authenticated:
        return redirect('/api/user_action?action=login')
    
    # 관리자만 접근 가능
    try:
        profile = UserProfile.objects.get(username=u.username)
        if not profile.is_admin:
            return redirect('/api/work')
    except:
        return redirect('/api/work')
    
    # 검색/필터 파라미터
    search = request.GET.get('search', '')
    selected_status = request.GET.get('status', '')
    page = request.GET.get('page', 1)
    
    # 토큰 목록 조회
    tokens = RustDeskToken.objects.all().order_by('-create_time')
    
    # 검색 필터
    if search:
        tokens = tokens.filter(
            Q(username__icontains=search) |
            Q(rid__icontains=search) |
            Q(uuid__icontains=search)
        )
    
    # 토큰 유효 시간 (2시간)
    active_threshold = datetime.datetime.now() - datetime.timedelta(seconds=7200)
    
    # 상태 필터
    if selected_status == 'active':
        tokens = tokens.filter(create_time__gte=active_threshold)
    elif selected_status == 'expired':
        tokens = tokens.filter(create_time__lt=active_threshold)
    
    # 통계
    all_tokens = RustDeskToken.objects.all()
    total_count = all_tokens.count()
    active_count = all_tokens.filter(create_time__gte=active_threshold).count()
    expired_count = total_count - active_count
    
    # 활성 상태 추가
    token_list = []
    for token in tokens:
        token.is_active = token.create_time >= active_threshold if token.create_time else False
        token_list.append(token)
    
    # 페이징
    paginator = Paginator(token_list, 50)
    page_obj = paginator.get_page(page)
    
    return render(request, 'token_manage.html', {
        'u': profile,
        'tokens': page_obj,
        'page_obj': page_obj,
        'total_count': total_count,
        'active_count': active_count,
        'expired_count': expired_count,
        'search': search,
        'selected_status': selected_status,
    })


def token_manage_delete(request, token_id):
    """토큰 삭제 API"""
    u = request.user
    if not u.is_authenticated:
        return JsonResponse({'code': 0, 'msg': '로그인이 필요합니다.'})
    
    try:
        profile = UserProfile.objects.get(username=u.username)
        if not profile.is_admin:
            return JsonResponse({'code': 0, 'msg': '권한이 없습니다.'})
    except:
        return JsonResponse({'code': 0, 'msg': '권한이 없습니다.'})
    
    try:
        token = RustDeskToken.objects.filter(id=token_id).first()
        if token:
            token.delete()
            return JsonResponse({'code': 1, 'msg': '삭제되었습니다.'})
        else:
            return JsonResponse({'code': 0, 'msg': '토큰을 찾을 수 없습니다.'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})


def peer_manage(request):
    """Peers 관리 페이지"""
    u = request.user
    if not u.is_authenticated:
        return redirect('/api/user_action?action=login')
    
    # 관리자만 접근 가능
    try:
        profile = UserProfile.objects.get(username=u.username)
        if not profile.is_admin:
            return redirect('/api/work')
    except:
        return redirect('/api/work')
    
    # 검색/필터 파라미터
    search = request.GET.get('search', '')
    selected_platform = request.GET.get('platform', '')
    page = request.GET.get('page', 1)
    
    # Peers 목록 조회
    peers = RustDeskPeer.objects.all().order_by('-rid')
    
    # 플랫폼 목록 (필터용)
    platforms = peers.values_list('platform', flat=True).distinct()
    platforms = [p for p in platforms if p]
    
    # 검색 필터
    if search:
        peers = peers.filter(
            Q(rid__icontains=search) |
            Q(uid__icontains=search) |
            Q(username__icontains=search) |
            Q(hostname__icontains=search) |
            Q(alias__icontains=search)
        )
    
    # 플랫폼 필터
    if selected_platform:
        peers = peers.filter(platform=selected_platform)
    
    # 통계
    total_count = RustDeskPeer.objects.count()
    
    # 태그 리스트 추가
    peer_list = []
    for peer in peers:
        peer.tags_list = [t.strip() for t in peer.tags.split(',') if t.strip()] if peer.tags else []
        peer_list.append(peer)
    
    # 페이징
    paginator = Paginator(peer_list, 50)
    page_obj = paginator.get_page(page)
    
    return render(request, 'peer_manage.html', {
        'u': profile,
        'peers': page_obj,
        'page_obj': page_obj,
        'total_count': total_count,
        'search': search,
        'platforms': platforms,
        'selected_platform': selected_platform,
    })


def peer_manage_delete(request, peer_id):
    """Peer 삭제 API"""
    u = request.user
    if not u.is_authenticated:
        return JsonResponse({'code': 0, 'msg': '로그인이 필요합니다.'})
    
    try:
        profile = UserProfile.objects.get(username=u.username)
        if not profile.is_admin:
            return JsonResponse({'code': 0, 'msg': '권한이 없습니다.'})
    except:
        return JsonResponse({'code': 0, 'msg': '권한이 없습니다.'})
    
    try:
        peer = RustDeskPeer.objects.filter(id=peer_id).first()
        if peer:
            peer.delete()
            return JsonResponse({'code': 1, 'msg': '삭제되었습니다.'})
        else:
            return JsonResponse({'code': 0, 'msg': 'Peer를 찾을 수 없습니다.'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})


@login_required(login_url='/api/user_action?action=login')
def user_bulk_template(request):
    """사용자 일괄 등록 엑셀 양식 다운로드"""
    username = request.user
    u = UserProfile.objects.get(username=username)
    
    # 최고관리자 또는 그룹관리자만 접근 가능
    if not (u.is_superuser or u.is_group_admin):
        return HttpResponse(_('권한이 없습니다.'), status=403)
    
    # 엑셀 파일 생성
    f = xlwt.Workbook(encoding='utf-8')
    sheet = f.add_sheet('사용자 일괄등록', cell_overwrite_ok=True)
    
    # 헤더 스타일
    header_style = xlwt.XFStyle()
    header_font = xlwt.Font()
    header_font.bold = True
    header_style.font = header_font
    
    # 필수 표시 스타일
    required_style = xlwt.XFStyle()
    required_font = xlwt.Font()
    required_font.bold = True
    required_font.colour_index = 2  # 빨간색
    required_style.font = required_font
    
    # 헤더 작성 (컬럼명)
    headers = [
        ('사용자명*', '필수, 4자 이상, 중복불가'),
        ('비밀번호*', '필수, 8~20자'),
        ('회사명', '선택'),
        ('이메일', '선택'),
        ('휴대전화', '선택'),
        ('회원등급', 'free/basic/standard/premium/enterprise (기본: free)'),
        ('최대상담원', '숫자 (기본: 3)'),
        ('활성화', 'Y/N (기본: Y)'),
        ('관리자', 'Y/N (기본: N)'),
    ]
    
    # 헤더 행
    for col, (header, desc) in enumerate(headers):
        if '*' in header:
            sheet.write(0, col, header, required_style)
        else:
            sheet.write(0, col, header, header_style)
    
    # 설명 행
    desc_style = xlwt.XFStyle()
    desc_font = xlwt.Font()
    desc_font.colour_index = 23  # 회색
    desc_style.font = desc_font
    
    for col, (header, desc) in enumerate(headers):
        sheet.write(1, col, desc, desc_style)
    
    # 예시 데이터 행
    example_data = [
        ['user001', 'password123', '테스트회사', 'user001@example.com', '010-1234-5678', 'free', '3', 'Y', 'N'],
        ['user002', 'password456', '샘플회사', 'user002@example.com', '010-9876-5432', 'basic', '5', 'Y', 'N'],
    ]
    
    for row, data in enumerate(example_data, start=2):
        for col, value in enumerate(data):
            sheet.write(row, col, value)
    
    # 컬럼 너비 조정
    col_widths = [15, 15, 20, 25, 15, 15, 12, 10, 10]
    for col, width in enumerate(col_widths):
        sheet.col(col).width = width * 256
    
    # 응답 생성
    sio = BytesIO()
    f.save(sio)
    sio.seek(0)
    
    response = HttpResponse(sio.getvalue(), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=user_bulk_template.xls'
    return response


@login_required(login_url='/api/user_action?action=login')
def user_bulk_upload(request):
    """사용자 일괄 등록 (엑셀 업로드)"""
    result = {'code': 0, 'msg': '', 'success_count': 0, 'fail_count': 0, 'errors': []}
    
    if request.method != 'POST':
        result['msg'] = _('잘못된 요청입니다.')
        return JsonResponse(result)
    
    username = request.user
    u = UserProfile.objects.get(username=username)
    
    # 최고관리자 또는 그룹관리자만 접근 가능
    if not (u.is_superuser or u.is_group_admin):
        result['msg'] = _('권한이 없습니다.')
        return JsonResponse(result)
    
    # 파일 확인
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        result['msg'] = _('엑셀 파일을 선택해주세요.')
        return JsonResponse(result)
    
    # 파일 확장자 확인
    file_name = excel_file.name.lower()
    if not (file_name.endswith('.xls') or file_name.endswith('.xlsx')):
        result['msg'] = _('엑셀 파일(.xls, .xlsx)만 업로드 가능합니다.')
        return JsonResponse(result)
    
    try:
        # 엑셀 파일 읽기
        import openpyxl
        from openpyxl import load_workbook
        
        # .xls 파일인 경우 xlrd 사용
        if file_name.endswith('.xls'):
            import xlrd
            workbook = xlrd.open_workbook(file_contents=excel_file.read())
            sheet = workbook.sheet_by_index(0)
            
            # 데이터 읽기 (3행부터 - 헤더와 설명 행 건너뛰기)
            rows = []
            for row_idx in range(2, sheet.nrows):  # 0: 헤더, 1: 설명, 2~: 데이터
                row_data = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    row_data.append(str(cell_value).strip() if cell_value else '')
                rows.append(row_data)
        else:
            # .xlsx 파일인 경우 openpyxl 사용
            workbook = load_workbook(filename=BytesIO(excel_file.read()))
            sheet = workbook.active
            
            # 데이터 읽기 (3행부터)
            rows = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                if row[0]:  # 첫 번째 컬럼(사용자명)이 있는 경우만
                    row_data = [str(cell).strip() if cell else '' for cell in row[:9]]
                    rows.append(row_data)
        
        if not rows:
            result['msg'] = _('등록할 사용자 데이터가 없습니다. (3행부터 데이터를 입력해주세요)')
            return JsonResponse(result)
        
        # 릴레이 서버 기본값 조회
        default_server = RelayServer.objects.filter(is_default=True).first()
        default_relay = default_server.server_address if default_server else '222.239.231.91'
        
        # 사용자 등록 처리
        success_count = 0
        fail_count = 0
        errors = []
        
        for row_num, row in enumerate(rows, start=3):
            try:
                # 컬럼 추출 (부족한 경우 기본값으로 채움)
                while len(row) < 9:
                    row.append('')
                
                target_username = row[0].strip()
                password = row[1].strip()
                company_name = row[2].strip() if len(row) > 2 else ''
                email = row[3].strip() if len(row) > 3 else ''
                phone = row[4].strip() if len(row) > 4 else ''
                membership_level = row[5].strip().lower() if len(row) > 5 and row[5] else 'free'
                max_agents_str = row[6].strip() if len(row) > 6 and row[6] else '3'
                is_active_str = row[7].strip().upper() if len(row) > 7 and row[7] else 'Y'
                is_admin_str = row[8].strip().upper() if len(row) > 8 and row[8] else 'N'
                
                # 필수 필드 검증
                if not target_username:
                    errors.append(f'{row_num}행: 사용자명이 비어있습니다.')
                    fail_count += 1
                    continue
                
                if len(target_username) <= 3:
                    errors.append(f'{row_num}행: 사용자명({target_username})은 3자 이상이어야 합니다.')
                    fail_count += 1
                    continue
                
                if not password:
                    errors.append(f'{row_num}행: 비밀번호가 비어있습니다.')
                    fail_count += 1
                    continue
                
                if len(password) < 8 or len(password) > 20:
                    errors.append(f'{row_num}행: 비밀번호({target_username})는 8~20자여야 합니다.')
                    fail_count += 1
                    continue
                
                # 사용자명 중복 확인
                if UserProfile.objects.filter(username=target_username).exists():
                    errors.append(f'{row_num}행: 사용자명({target_username})이 이미 존재합니다.')
                    fail_count += 1
                    continue
                
                # 회원등급 검증
                valid_levels = ['free', 'basic', 'standard', 'premium', 'enterprise']
                if membership_level not in valid_levels:
                    membership_level = 'free'
                
                # 최대 상담원 수 변환
                try:
                    max_agents = int(float(max_agents_str))
                    if max_agents < 1:
                        max_agents = 3
                except:
                    max_agents = 3
                
                # 활성화/관리자 여부 변환
                is_active = is_active_str in ['Y', 'YES', '1', 'TRUE', '예', '활성']
                is_admin = is_admin_str in ['Y', 'YES', '1', 'TRUE', '예', '관리자']
                
                # 그룹관리자인 경우 자신의 그룹으로만 저장
                if u.is_group_admin and not u.is_superuser:
                    # group_id가 있는 경우 같은 group_id로 저장
                    if u.group_id:
                        group_id = u.group_id
                        # 엑셀의 company_name은 무시하고 자신의 company_name 사용
                        company_name = u.company_name if u.company_name else company_name
                    # company_name이 있는 경우 같은 company_name으로 저장
                    elif u.company_name:
                        group_id = None
                        company_name = u.company_name
                    else:
                        errors.append(f'{row_num}행: 그룹관리자의 소속 정보가 없어 등록할 수 없습니다.')
                        fail_count += 1
                        continue
                    # 그룹관리자는 관리자 권한 부여 불가
                    is_admin = False
                else:
                    # 최고관리자는 엑셀의 group_id를 사용 (없으면 None)
                    group_id = None
                
                # 사용자 생성
                new_user = UserProfile(
                    username=target_username,
                    company_name=company_name,
                    group_id=group_id,
                    email=email,
                    phone=phone,
                    membership_level=membership_level,
                    max_agents=max_agents,
                    relay_server=default_relay,
                    is_active=is_active,
                    is_admin=is_admin,
                    is_group_admin=False  # 그룹관리자는 is_group_admin 설정 불가
                )
                new_user.set_password(password)
                new_user.save()
                
                success_count += 1
                
            except Exception as e:
                errors.append(f'{row_num}행: 처리 중 오류 발생 - {str(e)}')
                fail_count += 1
        
        result['code'] = 1 if success_count > 0 else 0
        result['success_count'] = success_count
        result['fail_count'] = fail_count
        result['errors'] = errors[:20]  # 최대 20개까지만 반환
        
        if success_count > 0 and fail_count == 0:
            result['msg'] = f'{success_count}명의 사용자가 성공적으로 등록되었습니다.'
        elif success_count > 0 and fail_count > 0:
            result['msg'] = f'{success_count}명 등록 성공, {fail_count}명 등록 실패'
        else:
            result['msg'] = f'등록에 실패했습니다. ({fail_count}건 오류)'
        
    except Exception as e:
        result['msg'] = _('엑셀 파일 처리 중 오류가 발생했습니다: {}').format(str(e))
        import traceback
        print(f'[엑셀 업로드 오류] {traceback.format_exc()}')
    
    return JsonResponse(result)
